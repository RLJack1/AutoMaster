import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import random
import math
from openpyxl import load_workbook
from PIL import Image, ImageTk
import os
import sys
import re
import warnings
import shutil

# Suppress openpyxl warnings for .xlsm files
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# =============================================================================
# MAINTAINER'S CHEAT SHEET — HOW TO HANDLE QCL COLUMN CHANGES
# =============================================================================
#
# AutoMaster reads data from:
#   1. QA Review file  — Excel workbook with sheets: Reopened Cases, Breached
#                        Cases, CSATs, CSAT Chat, Live Chat FCR, WD Breach
#   2. Member List     — Excel workbook listing team members and corp keys
#
# It then writes data into the QCL template (the output Excel file).
#
# There are TWO separate column systems you may need to change:
#   A) SOURCE columns  — columns READ from the QA Review file
#   B) QCL OUTPUT columns — columns WRITTEN to the QCL template
#
# ─────────────────────────────────────────────────────────────────────────────
# PART A: SOURCE FILE COLUMNS (Reading from QA Review file)
# ─────────────────────────────────────────────────────────────────────────────
#
# Source columns are found by searching for a keyword in the column header.
# The code uses this pattern:
#
#   some_col = find_column_containing(df, ["keyword"])
#
# Where:
#   - "df" is the sheet being read (e.g., reopened_df, csats_df)
#   - "keyword" is a partial match for the column header (case-insensitive)
#     It also works for headers like "F09 - SNOW HR Cases[keyword]"
#   - Multiple keywords can be listed as fallbacks: ["new name", "old name"]
#
#   TO RENAME A SOURCE COLUMN (header changed in the QA Review file):
#     Find the matching find_column_containing(...) line and update the keyword.
#     Example:
#       BEFORE:  some_col = find_column_containing(df, ["old header name"])
#       AFTER:   some_col = find_column_containing(df, ["new header name"])
#     Tip: List both for backwards compatibility:
#       some_col = find_column_containing(df, ["new header name", "old header name"])
#
#   TO ADD A NEW SOURCE COLUMN:
#     Add a new line in the "Column mapping" block of the relevant function:
#       new_col = find_column_containing(df, ["exact column header keyword"])
#     Then also write it to the QCL output (see Part B below).
#
#   TO REMOVE A SOURCE COLUMN:
#     Delete (or comment out) its find_column_containing(...) line.
#     Also delete its matching ws[f"X{current_row}"].value = ... line (Part B).
#
# ─────────────────────────────────────────────────────────────────────────────
# PART B: QCL OUTPUT COLUMNS (Writing to the QCL template)
# ─────────────────────────────────────────────────────────────────────────────
#
# QCL output columns are written using this pattern:
#
#   ws[f"X{current_row}"].value = variable
#
# Where:
#   - "X" is the LETTER of the column in the QCL Excel sheet (A, B, C, ...)
#   - "variable" holds the data to write into that cell
#
#   TO ADD A COLUMN to the QCL template:
#     Add a new line in the writing block of the relevant function:
#       ws[f"X{current_row}"].value = case[new_col]
#     Replace "X" with the correct QCL column letter (e.g., "L" for column L).
#     Make sure you also added the source column read (Part A) if needed.
#
#   TO REMOVE A COLUMN from the QCL template:
#     Delete (or comment out) the ws[f"X{current_row}"].value = ... line.
#     Also remove the source column read (Part A) if it is no longer used.
#
#   TO MOVE A COLUMN (column letter changed in the QCL template):
#     Update the letter in the ws[f"X{current_row}"].value line.
#     Example: if "Breach Reason" moves from column J to column K:
#       BEFORE:  ws[f"J{current_row}"].value = case[breach_reason_col]
#       AFTER:   ws[f"K{current_row}"].value = case[breach_reason_col]
#
# ─────────────────────────────────────────────────────────────────────────────
# SHEET REFERENCE — Which function handles which QCL sheet
# ─────────────────────────────────────────────────────────────────────────────
#
#   QCL Sheet            Function                               Applicable To
#   ────────────────     ─────────────────────────────────────  ─────────────
#   Reopened Cases       process_qa_reviews_reopened_cases()    All circles
#   Breached Cases       process_qa_reviews_breached_cases()    All circles
#   CSAT                 process_qa_reviews_csats()             All circles
#   CSAT Chat            process_qa_reviews_csat_chat()         Circle 6 only
#   Live Chat FCR        process_qa_reviews_live_chat_fcr()     Circle 6 only
#   WD Breached Cases    process_qa_reviews_wd_breach()         Katowice only
#
# Each function contains its own column map table and step-by-step notes
# directly above the column reading and writing code blocks.
#
# ─────────────────────────────────────────────────────────────────────────────
# PART C: CIRCLE AND TEAM MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────
#
# Teams are mapped to Circles in two places in this file:
#
#   1. get_circle_from_assignment_group()  — the core mapping function
#      Each elif block returns a circle number (1-6) based on keywords found
#      in the "Assignment Group" column of the source file.
#
#   2. The circle dropdown in the GUI (near the bottom of this file)
#      Lists the human-readable dropdown options shown to the user.
#
# Current Circle → Team mapping:
#   Circle 1: Human Capital Management (HCM), Org Management
#   Circle 2: Expense, Reporting
#   Circle 3: Learning
#   Circle 4: International Mobility
#   Circle 5: Performance & Rewards, Travel
#   Circle 6: Contact Center, Recruitment Admin
#
#   TO ADD A TEAM to an existing Circle:
#     Step 1 — In get_circle_from_assignment_group(), find the elif block for
#              the target circle and add the new team's keyword with "or":
#                elif "existing team" in group_lower or "new team" in group_lower:
#                    return X  (X = circle number)
#     Step 2 — In format_assignment_group(), add a normalization rule if the
#              raw team name needs to be displayed differently in the QCL:
#                elif "new team" in group:
#                    return "Formatted Team Name"
#     Step 3 — Update the dropdown label for that circle in the GUI section
#              (search for "circle_dropdown" near the bottom of this file).
#
#   TO REMOVE A TEAM from a Circle:
#     Step 1 — In get_circle_from_assignment_group(), delete the team's keyword
#              from its elif block (remove the "or 'team keyword' in group_lower"
#              portion, or delete the whole elif if it was the only team).
#     Step 2 — In format_assignment_group(), delete its normalization rule if present.
#     Step 3 — Update the dropdown label for that circle in the GUI section.
#
#   TO RENAME A TEAM (keyword changed in source file):
#     In get_circle_from_assignment_group(), update the old keyword string to
#     the new one in the relevant elif block.
#     In format_assignment_group(), update the matching condition and/or the
#     return value if the displayed name also changes.
#
#   TO TRANSFER A TEAM between Circles (e.g., Travel moved from Circle 2 → 5):
#     Step 1 — In get_circle_from_assignment_group(), remove the team's keyword
#              from the old Circle's elif block and add it to the new Circle's block.
#     Step 2 — Update both affected dropdown labels in the GUI section.
#     Example (Travel, Circle 2 → Circle 5):
#       BEFORE — Circle 2 block:  "expense" in group_lower or "travel" in group_lower or "reporting" in group_lower
#       AFTER  — Circle 2 block:  "expense" in group_lower or "reporting" in group_lower
#       AFTER  — Circle 5 block:  "performance" in group_lower or ... or "travel" in group_lower
#
#   TO ADD A NEW CIRCLE:
#     Step 1 — Add a new elif block in get_circle_from_assignment_group() with
#              the new circle number and its team keywords.
#     Step 2 — Add normalization rules in format_assignment_group() as needed.
#     Step 3 — Add the new circle as an option in the circle_dropdown in the GUI.
#
# =============================================================================

# Dictionary containing the country codes for all available countries serviced by CPS
COUNTRY_LOCATION_MAP = {
    "Albania": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "AL"
        ]
    },
    "American Samoa": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "AS"
        ]
    },
    "Andorra": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "AD"
        ]
    },
    "Antarctica": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "AQ"
        ]
    },
    "Australia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "SYD", "TUG", "WYN", "AU", "OVR", "Guest AU - Guest.AU", "MEL"
        ]
    },
    "Austria": {
        "formatted_prefixes": ["PH2_AT"],
        "arbitrary_codes": [
            "AT"
        ]
    },
    "Bangladesh": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "BD"
        ]
    },
    "Belarus": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "BY"
        ]
    },
    "Belgium": {
        "formatted_prefixes": ["PH3_BE_ING_BANK", "PH3_BE"],
        "arbitrary_codes": [
            "BE", "M1 03 14", "Guest BE - Guest.BE"
        ]
    },
    "Bhutan": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "BT"
        ]
    },
    "Bosnia and Herzegovina": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "BA"
        ]
    },
    "Brazil": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "Brazil"
        ]
    },
    "British Indian Ocean Territory": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "IO"
        ]
    },
    "Brunei Darussalam": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "BN"
        ]
    },
    "Bulgaria": {
        "formatted_prefixes": ["PH2_BG"],
        "arbitrary_codes": [
            "BG"
        ]
    },
    "Cambodia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "KH"
        ]
    },
    "China": {
        "formatted_prefixes": ["PH2_CN"],
        "arbitrary_codes": [
            "CN", "China"
        ]
    },
    "Christmas Island": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "CX"
        ]
    },
    "Cocos (Keeling) Islands": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "CC"
        ]
    },
    "Cook Islands": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "CK"
        ]
    },
    "Croatia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "HR"
        ]
    },
    "Cyprus": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "CY"
        ]
    },
    "Czech Republic": {
        "formatted_prefixes": ["PH2_CZ"],
        "arbitrary_codes": [
            "Cz"
        ]
    },
    "Denmark": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "DK"
        ]
    },
    "Estonia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "EE"
        ]
    },
    "Faroe Islands": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "FO"
        ]
    },
    "Fiji": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "FJ"
        ]
    },
    "Finland": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "FI"
        ]
    },
    "France": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "France", "RIC", "FR", "BER"
        ]
    },
    "French Polynesia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "PF"
        ]
    },
    "French Souther Territoties": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "TF"
        ]
    },
    "Georgia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "GE"
        ]
    },
    "Germany": {
        "formatted_prefixes": ["PH2_DE"],
        "arbitrary_codes": [
            "DE", "Germany", "Guest DE - Guest.DE"
        ]
    },
    "Gibraltar": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "GI"
        ]
    },
    "Greece": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "GR"
        ]
    },
    "Guam": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "GU"
        ]
    },
    "Heard Island and McDonald Islands": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "HM"
        ]
    },
    "Hong Kong": {
        "formatted_prefixes": ["PH2_HK"],
        "arbitrary_codes": [
            "Hong Kong", "HK", "Hk"
        ]
    },
    "Hungary": {
        "formatted_prefixes": ["PH2_HU"],
        "arbitrary_codes": [
            "HU"
        ]
    },
    "Iceland": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "IS"
        ]
    },
    "India": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "IN"
        ]
    },
    "Indonesia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "ID"
        ]
    },
    "Ireland": {
        "formatted_prefixes": ["PH2_IE"],
        "arbitrary_codes": [
            "IE"
        ]
    },
    "Isle of Man": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "IM"
        ]
    },
    "Italy": {
        "formatted_prefixes": ["PH2_ITA", "PH3_ITA"],
        "arbitrary_codes": [
            "Italy", "IT"
        ]
    },
    "Japan": {
        "formatted_prefixes": ["PH2_JP"],
        "arbitrary_codes": [
            "Japan", "JP"
        ]
    },
    "Kiribati": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "KI"
        ]
    },
    "Kosovo": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "XK"
        ]
    },
    "Kyrgyzstan": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "KG"
        ]
    },
    "Laos": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "LA"
        ]
    },
    "Latvia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "LV"
        ]
    },
    "Liechtenstein": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "LI"
        ]
    },
    "Lithuania": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "LT"
        ]
    },
    "Luxembourg": {
        "formatted_prefixes": ["PH2_LU"],
        "arbitrary_codes": [
            "LU"
        ]
    },
    "Macao": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MO"
        ]
    },
    "Macedonia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MK"
        ]
    },
    "Malaysia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MY"
        ]
    },
    "Maldives": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MV"
        ]
    },
    "Malta": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MT"
        ]
    },
    "Marshall Islands": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MH"
        ]
    },
    "Mexico": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MX"
        ]
    },
    "Micronesia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "FM"
        ]
    },
    "Moldova": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MD"
        ]
    },
    "Monaco": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MC"
        ]
    },
    "Mongolia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MN"
        ]
    },
    "Montenegro": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "ME"
        ]
    },
    "Myanmar": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MM"
        ]
    },
    "Nauru": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "NR"
        ]
    },
    "Nepal": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "NP"
        ]
    },
    "Netherlands": {
        "formatted_prefixes": ["Guest NL"],
        "arbitrary_codes": [
            "ACT", "ALP", "AME",
            "BMG", 
            "CDR", 
            "DEA", "DP", 
            "HBP",
            "KBK","KCC", "KCM", "KFL", "KFR", "KGD", "KMH", "KNF", "KQB", "KQB", "KSY", "KXS", "KZK", "KZR", "KNZ", "KBE", "KLQ", "KQT", "KZW", "KNK", "KRQ",
            "LZ",
            "NL",
            "RBG", 
            "WP","WBM","WCD",
            "Guest NL - Guest.NL",
        ]
    },
    "New Caledonia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "NC"
        ]
    },
    "New Zealand": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "NZ"
        ]
    },
    "Niue": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "NU"
        ]
    },
    "Norfolk Island": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "NF"
        ]
    },
    "Northern Mariana Islands": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "MP"
        ]
    },
    "Norway": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "NO"
        ]
    },
    "Pakistan": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "PK"
        ]
    },
    "Palau": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "PW"
        ]
    },
    "Papua New Guinea": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "PG"
        ]
    },
    "Philippines": {
        "formatted_prefixes": ["PH1_PH", "Guest PH", "PH2_PH"],
        "arbitrary_codes": [
            "PH", "Guest PH - Guest.PH"
        ]
    },
    "Philippines Bank": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "PH Bank", "Guest PHB - Guest.PHB", "PHB"
        ]
    },
    "Pitcairn": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "PN"
        ]
    },
    "Poland": {
        "formatted_prefixes": ["PH1_PL"],
        "arbitrary_codes": [
            "PL", "PH2_SK_Brati", "P-PULAWSKA", "P-CHORZOWSKA", "P-SOKOLSKA", "P-CHORZ.50", "DR01R1402", "DR09R0102"
            "DR02R0602", "DR06R0701R"
        ]
    },
    "Poland Bank": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "PL Bank"
        ]
    },
    "Portugal": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "PT",
        ]
    },
    "Romania": {
        "formatted_prefixes": ["PH1_RO"],
        "arbitrary_codes": [
            "RO"
        ]
    },
    "Russian Federation": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "RUSMCW001", "RU"
        ]
    },
    "Samoa": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "WS"
        ]
    },
    "San Marino": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "SM"
        ]
    },
    "Serbia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "RS"
        ]
    },
    "Singapore": {
        "formatted_prefixes": ["PH2_SG"],
        "arbitrary_codes": [
            "SG", "Guest Asia - Guest.Asia", "Guest ASIA - Guest.ASIA", "ASIA"
        ]
    },
    "Slovakia": {
        "formatted_prefixes": ["PH1_SK","PH2_SK"],
        "arbitrary_codes": [
            "SK",
        ]
    },
    "Slovenia": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "SI"
        ]
    },
    "Solomon Islands": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "SB"
        ]
    },
    "South Korea": {
        "formatted_prefixes": ["PH2_KR"],
        "arbitrary_codes": [
            "Kr", "KR"
        ]
    },
    "South Sudan": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "SS"
        ]
    },
    "Spain": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "Spain", "Madrid_Pobla", "Madrid_Poblados", "Valladolid", "ES_Madrid_Hubs"
        ]
    },
    "Sri Lanka": {
        "formatted_prefixes": ["PH2_LK"],
        "arbitrary_codes": [
        ]
    },
    "Sweden": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "SE"
        ]
    },
    "Switzerland": {
        "formatted_prefixes": ["PH2_CH"],
        "arbitrary_codes": [
            "CH"
        ]
    },
    "Taiwan": {
        "formatted_prefixes": ["PH2_TW"],
        "arbitrary_codes": [
            "TW"
        ]
    },
    "Tajikistan": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "TJ"
        ]
    },
    "Thailand": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "TH"
        ]
    },
    "Timor-Leste": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "TL"
        ]
    },
    "Tokelau": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "TK"
        ]
    },
    "Tonga": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "TO"
        ]
    },
    "Turkey": {
        "formatted_prefixes": ["PH2_TR"],
        "arbitrary_codes": [
            "TR", "4127", "001001", "001015"
        ]
    },
    "Turkmenistan": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "TM"
        ]
    },
    "Tuvalu": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "TV"
        ]
    },
    "Ukraine": {
        "formatted_prefixes": ["PH2_UA"],
        "arbitrary_codes": [
            "UA"
        ]
    },
    "United Arab Emirates": {
        "formatted_prefixes": ["PH2_AE"],
        "arbitrary_codes": [
        ]
    },
    "United Kingdom": {
        "formatted_prefixes": ["PH2_GB"],
        "arbitrary_codes": [
            "Guest UK - Guest.UK", "United Kingdom", "UK"
        ]
    },
    "United States": {
        "formatted_prefixes": ["PH2_US"],
        "arbitrary_codes": [
        ]
    },
    "United States Minor Outlying Islands": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "UM"
        ]
    },
    "Uzbekistan": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "UZ"
        ]
    },
    "Vanuatu": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "VU"
        ]
    },
    "Vatican City State": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "VA"
        ]
    },
    "Vietnam": {
        "formatted_prefixes": ["PH2_VN"],
        "arbitrary_codes": [
            "Vn", "VN"
        ]
    },
    "Wallis and Futuna": {
        "formatted_prefixes": [],
        "arbitrary_codes": [
            "WF"
        ]
    },
}

# ---------------- HELPER FUNCTIONS ---------------- 

def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def load_image(image_path, width=None, height=None):
    """Load and resize image for tkinter"""
    try:
        img = Image.open(get_resource_path(image_path))
        if width and height:
            img = img.resize((width, height), Image.Resampling.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception as e:
        print(f"Error loading image {image_path}: {e}")
        return None

# ---------------- HELPER FUNCTIONS ---------------- 

def normalize_raw_name(raw_name):
    """
    Converts various name formats to 'Last, First'
    Handles: 'First Last - ID', 'First Last', 'Last, First'
    """
    if pd.isna(raw_name):
        return None
    
    raw_name = str(raw_name).strip()
    
    # Handle empty strings
    if not raw_name:
        return None
    
    # Remove ID if present (e.g., "First Last - ID123")
    if "-" in raw_name:
        raw_name = raw_name.split("-")[0].strip()
    
    # If already in "Last, First" format, return as-is
    if "," in raw_name:
        return raw_name.strip()
    
    # Convert "First Last" to "Last, First"
    parts = raw_name.split()
    if len(parts) >= 2:
        first = parts[0]
        last = " ".join(parts[1:])  # Handle multi-part last names
        return f"{last}, {first}"
    
    # If single name, return as-is
    return raw_name

def format_assignment_group(group):
    if pd.isna(group):
        return ""
    
    # Switch-case style normalization for Assignment Group
    
    group = str(group).lower()
    if "compensation" in group or "benefits" in group:
        return "Performance & Rewards"
    elif "expenses" in group:
        return "Expenses"
    elif "hcm" in group:
        return "Human Capital Management"
    elif "organizational management" in group:
        return "Org Management"
    elif "international mobility" in group:
        return "International Mobility"
    elif "learning coordination" in group:
        return "Learning"
    elif "people services" in group:
        return "Contact Center"
    elif "recruitment admin" in group:
        return "Recruitment Admin"
    elif "reporting" in group:
        return "Reporting"
    elif "travels" in group:
        return "Travel"
    else:
        return group.title()

def calculate_csat_type(csat_score):
    """
    Calculates CSAT Type based on CSAT Score
    1-2 = Dissatisfied
    3 = Neutral
    4-5 = Satisfied
    """
    if pd.isna(csat_score):
        return ""
    
    try:
        score = int(csat_score)
        if score >= 1 and score <= 2:
            return "Dissatisfied"
        elif score == 3:
            return "Neutral"
        elif score >= 4 and score <= 5:
            return "Satisfied"
        else:
            return ""
    except:
        return ""

def calculate_response_category(comment):
    """
    Calculates Response Category based on Comments
    If comment is empty/blank = "No comment"
    Otherwise = leave blank
    """
    if pd.isna(comment) or str(comment).strip() == "":
        return "No comment"
    else:
        return ""

def calculate_sample_percentage(tenure):
    if tenure < 6:
        return 0.15 / 2
    elif tenure < 12:
        return 0.10 / 2
    else:
        return 0.05 / 2
    
def extract_formatted_prefix(location_code):
    """
    Extracts Department-Country prefix from formatted codes.
    Example: PH1_BE_Brussels -> PH1_BE
    """
    parts = location_code.split("_")
    if len(parts) >= 2:
        return f"{parts[0]}_{parts[1]}"
    return location_code


def format_location(raw_location):
    """
    Returns a tuple: (formatted_location, was_recognized, raw_value)
    - was_recognized is True if location matched the dictionary
    - was_recognized is False if location was unrecognized
    - was_recognized is None if location was blank/empty
    - raw_value is the original input (for logging unrecognized codes)
    """
    if pd.isna(raw_location) or str(raw_location).strip() == "":
        return ("", None, "")  # Blank location

    raw_location_str = str(raw_location).strip()
    raw_location_upper = raw_location_str.upper()

    # Extract formatted prefix if applicable (only for formatted prefix checking)
    prefix = extract_formatted_prefix(raw_location_upper)

    # FIRST PASS: Check formatted prefixes for ALL countries
    for country, rules in COUNTRY_LOCATION_MAP.items():
        if prefix in rules.get("formatted_prefixes", []):
            return (country, True, raw_location_str)

    # SECOND PASS: Check arbitrary codes with EXACT matching
    for country, rules in COUNTRY_LOCATION_MAP.items():
        for code in rules.get("arbitrary_codes", []):
            # Exact match only (case-insensitive)
            if code.upper() == raw_location_upper:
                return (country, True, raw_location_str)

    # No match found - return original value (unrecognized)
    return (raw_location_str, False, raw_location_str)

def infer_location_from_opened_for(opened_for_value):
    """
    Infers location from 'Opened for' field by checking for guest account patterns
    like "Guest XX - Guest.XX" where XX is a country code.
    Only matches codes that appear as whole words to avoid false positives
    (e.g. "GU" matching inside "Guest").
    Returns the country name if found, empty string otherwise.
    """
    if pd.isna(opened_for_value):
        return ""

    opened_for_str = str(opened_for_value).strip()
    opened_for_upper = opened_for_str.upper()

    # Split the opened_for value into distinct tokens for exact word matching
    # This handles patterns like "Guest NL - Guest.NL" by splitting on spaces, dots, and hyphens
    tokens = set(re.split(r'[\s\.\-]+', opened_for_upper))

    # Check against all arbitrary codes in the country location map
    for country, rules in COUNTRY_LOCATION_MAP.items():
        for code in rules.get("arbitrary_codes", []):
            if code.upper() in tokens:
                return country

    return ""

def get_case_location(location_value, opened_for_value):
    """
    Gets the location for a case, trying Location first, then falling back to Opened for.
    Returns a tuple: (formatted_location, was_recognized, raw_value)
    - was_recognized is True if location matched the dictionary
    - was_recognized is False if location was unrecognized
    - was_recognized is None if location was blank/empty
    - raw_value is the original location value (for logging unrecognized codes)
    """
    # Try to get location from Location column first
    location, recognized, raw_value = format_location(location_value)

    # If location is blank or unrecognized, try to infer from Opened for
    if recognized is None or recognized is False:
        inferred = infer_location_from_opened_for(opened_for_value)
        if inferred:
            return (inferred, True, raw_value)

    return (location, recognized, raw_value)

def find_column(df, possible_names):
    # Finds a column in df whose header matches one of the possible_names.
    # Matching is case-insensitive.
    
    for col in df.columns:
        col_name = str(col).strip().lower()
        for name in possible_names:
            if col_name == name.strip().lower():
                return col
    raise ValueError(f"Missing required column: {possible_names}")

def find_qcl_start_row(ws, search_column="B", search_text="control check no."):

    # Finds the first empty row below a specified search text in a specified column

    for row in range(1, ws.max_row + 1):
        cell_value = ws[f"{search_column}{row}"].value
        if cell_value and str(cell_value).strip().lower() == search_text.lower():
            r = row + 1
            while ws[f"{search_column}{r}"].value:
                r += 1
            return r
    raise ValueError(f"Could not find '{search_text}' in column {search_column} of QCL template.")

def detect_header_row(file_path, sheet_name=0, required_columns=None):

    # Detects the row number containing required column headers.
    # Returns row index to be used as header=.

    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    required_columns = [col.lower() for col in required_columns]
    for idx, row in preview.iterrows():
        row_values = [str(cell).strip().lower() for cell in row.values if pd.notna(cell)]
        if all(req in row_values for req in required_columns):
            return idx
    raise ValueError(f"Could not find header row containing: {required_columns}")

def extract_corpkey_from_name(formatted_name):

    # Extracts corp key from "FirstName LastName - CorpKey" format
    # Returns the corp key or None if not found

    if pd.isna(formatted_name):
        return None

    formatted_name = str(formatted_name).strip()

    # Check if it contains " - " separator
    if " - " in formatted_name:
        parts = formatted_name.split(" - ")
        if len(parts) >= 2:
            return parts[-1].strip()  # Return the last part (corp key)

    return None

def get_circle_from_assignment_group(assignment_group):

    # Maps assignment group to circle number (1-6)
    # Circle 1: Human Capital Management and Org Management
    # Circle 2: Expense and Reporting
    # Circle 3: Learning
    # Circle 4: International Mobility
    # Circle 5: Performance & Rewards and Travel
    # Circle 6: Contact Center and Recruitment Admin

    if pd.isna(assignment_group):
        return None

    group_lower = str(assignment_group).lower()

    # Circle 1: HCM and Org Management
    if "human capital management" in group_lower or "hcm" in group_lower or "org management" in group_lower or "organizational management" in group_lower:
        return 1
    # Circle 2: Expense and Reporting
    elif "expense" in group_lower or "reporting" in group_lower:
        return 2
    # Circle 3: Learning
    elif "learning" in group_lower:
        return 3
    # Circle 4: International Mobility
    elif "international mobility" in group_lower or "mobility" in group_lower:
        return 4
    # Circle 5: Performance & Rewards and Travel
    elif "performance" in group_lower or "rewards" in group_lower or "compensation" in group_lower or "benefits" in group_lower or "travel" in group_lower:
        return 5
    # Circle 6: Contact Center and Recruitment Admin
    elif "contact center" in group_lower or "people services" in group_lower or "recruitment admin" in group_lower:
        return 6
    else:
        return None

KATOWICE_OPTIONS = [
    "All Katowice",
    "HCM - BE",
    "HCM - NL",
    "People Services - DE",
    "People Services - NL",
    "People Services - PL"
]

# Maps Katowice dropdown options to the assignment group substrings in the raw file
KATOWICE_TEAM_MAP = {
    "HCM - BE": "PL CPS HCM - BE ING",
    "HCM - NL": "PL CPS HCM - NL ING",
    "People Services - DE": "PL CPS People Services - DE ING",
    "People Services - NL": "PL CPS People Services - NL ING",
    "People Services - PL": "PL CPS People Services - PL ING"
}

def is_katowice_selection(circle_selection):
    # Returns True if the dropdown selection is a Katowice option
    return circle_selection in KATOWICE_OPTIONS

def parse_circle_selection(circle_selection):

    # Parses the circle selection from the dropdown
    # Returns the circle number (1-6) for Manila, the raw string for Katowice, or None for "All Circles"

    if circle_selection == "All Circles":
        return None

    # Katowice options return the raw string
    if is_katowice_selection(circle_selection):
        return circle_selection

    # Extract circle number from strings like "Circle 1 (HCM, OM)"
    if "Circle" in circle_selection:
        try:
            circle_num = int(circle_selection.split()[1])
            return circle_num
        except:
            return None

    return None

# ---------------- MAIN PROCESS ---------------- #

def process_qa_reviews_reopened_cases(qa_review_file, member_file, wb, control_no_start, selected_circle=None, log_func=None, is_katowice=False):

    # Process QA Reviews - Reopened Cases sheet
    # selected_circle: None for all circles, or 1-6 for specific circle
    # log_func: Function to log messages to console
    # Returns: (next_control_no, total_cases_written)

    try:
        # Read the "Reopened Cases" sheet from QA Review file
        reopened_df = pd.read_excel(qa_review_file, sheet_name="Reopened Cases")

        # Find required columns in Reopened Cases sheet
        # All columns follow format: "F09 - SNOW HR Cases[Column Name]"
        # We'll search for columns that contain the target name within brackets
        def find_column_containing(df, search_terms):
            """Find column that contains any of the search terms (case-insensitive)"""
            for col in df.columns:
                col_lower = str(col).strip().lower()
                for term in search_terms:
                    # Check if the term is in the column name (handles both bracket and non-bracket formats)
                    if f"[{term.strip().lower()}]" in col_lower or term.strip().lower() in col_lower:
                        return col
            raise ValueError(f"Missing required column containing: {search_terms}")

        # ── COLUMN MAP: Reopened Cases ────────────────────────────────────────────
        # SOURCE KEYWORD (in QA Review file)          QCL OUTPUT COLUMN
        #   "assignment group"                     →  Column C  (Assignment Group)
        #   "country code"                         →  Column D  (Location)
        #   "user id"                              →  Column E  (Processor Name)
        #   "number"                               →  Column F  (Reference Number)
        #   "hr service"                           →  Column G  (HR Service)
        #   "re-opened reason" / "reopened reason" →  Column I  (Reopened Reason)
        #   Note: Column H is reserved in the QCL template — do not write to it.
        #
        # TO ADD A COLUMN:
        #   Step 1 — Add a read line below:
        #     new_col = find_column_containing(reopened_df, ["exact header keyword"])
        #   Step 2 — Add a write line in the WRITING BLOCK further below:
        #     ws[f"X{current_row}"].value = case[new_col]
        #     (Replace "X" with the target QCL column letter, e.g. "J" for column J)
        #
        # TO REMOVE A COLUMN:
        #   Step 1 — Delete (or comment out) its find_column_containing(...) line below.
        #   Step 2 — Delete (or comment out) its ws[f"X{current_row}"].value = ... line
        #            in the WRITING BLOCK further below.
        #
        # TO RENAME A SOURCE COLUMN (QA Review file header changed):
        #   Update the keyword in find_column_containing(reopened_df, ["old name"])
        #   to the new name. Tip: list both as fallbacks: ["new name", "old name"]
        #
        # TO MOVE A QCL COLUMN (column letter changed in QCL template):
        #   In the WRITING BLOCK, change ws[f"OLD_LETTER{current_row}"]
        #   to ws[f"NEW_LETTER{current_row}"]
        # ─────────────────────────────────────────────────────────────────────────
        assignment_group_col = find_column_containing(reopened_df, ["assignment group"])                    # → QCL Column C
        country_code_col = find_column_containing(reopened_df, ["country code"])                            # → QCL Column D
        user_id_col = find_column_containing(reopened_df, ["user id"])                                      # → QCL Column E
        number_col = find_column_containing(reopened_df, ["number"])                                        # → QCL Column F
        hr_service_col = find_column_containing(reopened_df, ["hr service"])                                # → QCL Column G
        reopened_reason_col = find_column_containing(reopened_df, ["re-opened reason", "reopened reason"])  # → QCL Column I

        # Read member list file
        member_header_row = detect_header_row(
            member_file,
            required_columns=["Team", "Tenure"]
        )
        members_df = pd.read_excel(member_file, header=member_header_row)

        # Find formatted name column (with corp key) and Team column (LastName, FirstName format)
        formatted_name_col = None
        for col in members_df.columns:
            sample_values = members_df[col].dropna().astype(str).head(5)
            if any(" - " in val for val in sample_values):
                formatted_name_col = col
                break

        member_name_col = formatted_name_col if formatted_name_col else find_column(members_df, ["Team"])
        team_col = find_column(members_df, ["Team"])  # Column A with "LastName, FirstName" format
        tenure_col = find_column(members_df, ["Tenure"])

        # Build corpkey to member mapping
        corpkey_to_member = {}
        for _, member in members_df.iterrows():
            member_name_raw = str(member[member_name_col]).strip()
            team_name = str(member[team_col]).strip()  # Get the "LastName, FirstName" format
            tenure = member[tenure_col]

            if not member_name_raw or member_name_raw.lower() == 'nan' or pd.isna(tenure):
                continue

            corpkey = extract_corpkey_from_name(member_name_raw)
            if corpkey:
                corpkey_to_member[corpkey] = {
                    'name': member_name_raw,
                    'team_name': team_name,  # Store the Team column name for display
                    'tenure': tenure,
                    'cases': []
                }

        # Get raw file name for logging
        raw_file_name = os.path.basename(qa_review_file)

        # Build dictionary of cases per corpkey with circle filtering (with row index)
        for row_idx, row in reopened_df.iterrows():
            user_id = str(row[user_id_col]).strip()
            if user_id and user_id.lower() != 'nan':
                # Check if case matches the selected circle/team filter
                if is_katowice:
                    if selected_circle == "All Katowice":
                        include_case = True
                    else:
                        ktw_team = KATOWICE_TEAM_MAP.get(selected_circle, "").lower()
                        include_case = ktw_team in str(row[assignment_group_col]).strip().lower()
                else:
                    case_circle = get_circle_from_assignment_group(row[assignment_group_col])
                    include_case = selected_circle is None or case_circle == selected_circle
                if include_case:
                    if user_id in corpkey_to_member:
                        corpkey_to_member[user_id]['cases'].append({'row': row, 'raw_row_num': row_idx + 2})

        # Access the Reopened Cases sheet
        if "Reopened Cases" not in wb.sheetnames:
            raise ValueError("QCL template does not contain 'Reopened Cases' sheet")

        ws = wb["Reopened Cases"]
        current_row = find_qcl_start_row(ws, search_column="B", search_text="control check no.")
        control_no = control_no_start

        # Track statistics
        total_cases_written = 0

        print(f"=== PROCESSING REOPENED CASES ===")

        # Process each member with cases
        for corpkey, member_data in corpkey_to_member.items():
            cases = member_data['cases']

            if len(cases) == 0:
                continue

            # Write ALL cases to QCL - Reopened Cases sheet
            for case_data in cases:
                case = case_data['row']
                raw_row_num = case_data['raw_row_num']
                case_number = case[number_col]
                location, recognized, raw_location = format_location(case[country_code_col])

                # ── WRITING BLOCK: Reopened Cases ────────────────────────────────────────
                # Each line writes one value into the QCL. The letter (B, C, D...) is the
                # column in the QCL Excel sheet. To add a new column here, add a new line
                # following the same pattern and use the correct QCL column letter.
                ws[f"B{current_row}"].value = control_no                                                                             # Col B: Control Check No. (auto-numbered)
                ws[f"C{current_row}"].value = str(case[assignment_group_col]).strip() if is_katowice else format_assignment_group(case[assignment_group_col])  # Col C: Assignment Group
                ws[f"D{current_row}"].value = location                                                                               # Col D: Location
                ws[f"E{current_row}"].value = member_data['name'] if is_katowice else member_data['team_name']                       # Col E: Processor Name
                ws[f"F{current_row}"].value = case_number                                                                            # Col F: Reference Number
                ws[f"G{current_row}"].value = case[hr_service_col]                                                                   # Col G: HR Service
                ws[f"I{current_row}"].value = case[reopened_reason_col]                                                              # Col I: Reopened Reason
                # ── To add a column: ws[f"X{current_row}"].value = case[new_col]
                #    (replace X with the QCL column letter, e.g. "J" for column J)
                # ─────────────────────────────────────────────────────────────────────────

                # Log statistics to terminal, unprocessed notices to GUI console
                if recognized is True:
                    print(f"Reopened Cases {control_no}: Case {case_number} from {location} was taken from row {raw_row_num} in {raw_file_name}/Reopened Cases")
                elif recognized is False:
                    if log_func:
                        log_func(f"Reopened Cases {control_no}: Case {case_number} has an unrecognized location code '{raw_location}'. Taken from row {raw_row_num} in {raw_file_name}/Reopened Cases.")
                else:  # recognized is None (blank)
                    if log_func:
                        log_func(f"Reopened Cases {control_no}: Case {case_number} has a blank location. Taken from row {raw_row_num} in {raw_file_name}/Reopened Cases.")

                control_no += 1
                current_row += 1
                total_cases_written += 1

        if log_func:
            log_func(f"Reopened Cases completed: {total_cases_written} cases written.\n")

        return control_no, total_cases_written

    except Exception as e:
        raise Exception(f"Error processing QA Reviews - Reopened Cases: {str(e)}")

def process_qa_reviews_breached_cases(qa_review_file, member_file, wb, control_no_start, selected_circle=None, log_func=None, is_katowice=False):

    # Process QA Reviews - Breached Cases sheet
    # selected_circle: None for all circles, or 1-6 for specific circle
    # log_func: Function to log messages to console
    # Returns: (next_control_no, total_cases_written)

    try:
        # Read the "Breached Cases" sheet from QA Review file
        breached_df = pd.read_excel(qa_review_file, sheet_name="Breached Cases")

        # Find required columns in Breached Cases sheet
        def find_column_containing(df, search_terms):
            """Find column that contains any of the search terms (case-insensitive)"""
            for col in df.columns:
                col_lower = str(col).strip().lower()
                for term in search_terms:
                    # Check if the term is in the column name (handles both bracket and non-bracket formats)
                    if f"[{term.strip().lower()}]" in col_lower or term.strip().lower() in col_lower:
                        return col
            raise ValueError(f"Missing required column containing: {search_terms}")

        # ── COLUMN MAP: Breached Cases ────────────────────────────────────────────
        # SOURCE KEYWORD (in QA Review file)    QCL OUTPUT COLUMN
        #   "assignment group"              →  Column C  (Assignment Group)
        #   "country code"                  →  Column D  (Location)
        #   "user id"                       →  Column E  (Processor Name)
        #   "number"                        →  Column F  (Reference Number)
        #   "hr service"                    →  Column G  (HR Service)
        #   "resolution type"               →  Column H  (SLA Breach Type)
        #   "sla breach reason"             →  Column J  (Breach Reason)
        #   Note: Column I is reserved in the QCL template — do not write to it.
        #
        # TO ADD A COLUMN:
        #   Step 1 — Add a read line below:
        #     new_col = find_column_containing(breached_df, ["exact header keyword"])
        #   Step 2 — Add a write line in the WRITING BLOCK further below:
        #     ws[f"X{current_row}"].value = case[new_col]
        #     (Replace "X" with the target QCL column letter, e.g. "K" for column K)
        #
        # TO REMOVE A COLUMN:
        #   Step 1 — Delete (or comment out) its find_column_containing(...) line below.
        #   Step 2 — Delete (or comment out) its ws[f"X{current_row}"].value = ... line
        #            in the WRITING BLOCK further below.
        #
        # TO RENAME A SOURCE COLUMN (QA Review file header changed):
        #   Update the keyword in find_column_containing(breached_df, ["old name"])
        #   to the new name. Tip: list both as fallbacks: ["new name", "old name"]
        #
        # TO MOVE A QCL COLUMN (column letter changed in QCL template):
        #   In the WRITING BLOCK, change ws[f"OLD_LETTER{current_row}"]
        #   to ws[f"NEW_LETTER{current_row}"]
        # ─────────────────────────────────────────────────────────────────────────
        assignment_group_col = find_column_containing(breached_df, ["assignment group"])  # → QCL Column C
        country_code_col = find_column_containing(breached_df, ["country code"])          # → QCL Column D
        user_id_col = find_column_containing(breached_df, ["user id"])                    # → QCL Column E
        number_col = find_column_containing(breached_df, ["number"])                      # → QCL Column F
        hr_service_col = find_column_containing(breached_df, ["hr service"])              # → QCL Column G
        resolution_type_col = find_column_containing(breached_df, ["resolution type"])    # → QCL Column H
        breach_reason_col = find_column_containing(breached_df, ["sla breach reason"])    # → QCL Column J

        # Read member list file
        member_header_row = detect_header_row(
            member_file,
            required_columns=["Team", "Tenure"]
        )
        members_df = pd.read_excel(member_file, header=member_header_row)

        # Find formatted name column (with corp key) and Team column (LastName, FirstName format)
        formatted_name_col = None
        for col in members_df.columns:
            sample_values = members_df[col].dropna().astype(str).head(5)
            if any(" - " in val for val in sample_values):
                formatted_name_col = col
                break

        member_name_col = formatted_name_col if formatted_name_col else find_column(members_df, ["Team"])
        team_col = find_column(members_df, ["Team"])  # Column A with "LastName, FirstName" format
        tenure_col = find_column(members_df, ["Tenure"])

        # Build corpkey to member mapping
        corpkey_to_member = {}
        for _, member in members_df.iterrows():
            member_name_raw = str(member[member_name_col]).strip()
            team_name = str(member[team_col]).strip()  # Get the "LastName, FirstName" format
            tenure = member[tenure_col]

            if not member_name_raw or member_name_raw.lower() == 'nan' or pd.isna(tenure):
                continue

            corpkey = extract_corpkey_from_name(member_name_raw)
            if corpkey:
                corpkey_to_member[corpkey] = {
                    'name': member_name_raw,
                    'team_name': team_name,  # Store the Team column name for display
                    'tenure': tenure,
                    'cases': []
                }

        # Get raw file name for logging
        raw_file_name = os.path.basename(qa_review_file)

        # Build dictionary of cases per corpkey with circle filtering (with row index)
        for row_idx, row in breached_df.iterrows():
            user_id = str(row[user_id_col]).strip()
            if user_id and user_id.lower() != 'nan':
                # Check if case matches the selected circle/team filter
                if is_katowice:
                    if selected_circle == "All Katowice":
                        include_case = True
                    else:
                        ktw_team = KATOWICE_TEAM_MAP.get(selected_circle, "").lower()
                        include_case = ktw_team in str(row[assignment_group_col]).strip().lower()
                else:
                    case_circle = get_circle_from_assignment_group(row[assignment_group_col])
                    include_case = selected_circle is None or case_circle == selected_circle
                if include_case:
                    if user_id in corpkey_to_member:
                        corpkey_to_member[user_id]['cases'].append({'row': row, 'raw_row_num': row_idx + 2})

        # Access the Breached Cases sheet
        if "Breached Cases" not in wb.sheetnames:
            raise ValueError("QCL template does not contain 'Breached Cases' sheet")

        ws = wb["Breached Cases"]
        current_row = find_qcl_start_row(ws, search_column="B", search_text="control check no.")
        control_no = control_no_start

        # Track statistics
        total_cases_written = 0

        print(f"=== PROCESSING BREACHED CASES ===")

        # Process each member with cases
        for corpkey, member_data in corpkey_to_member.items():
            cases = member_data['cases']

            if len(cases) == 0:
                continue

            # Write ALL cases to QCL - Breached Cases sheet (no sampling)
            for case_data in cases:
                case = case_data['row']
                raw_row_num = case_data['raw_row_num']
                case_number = case[number_col]
                location, recognized, raw_location = format_location(case[country_code_col])

                # ── WRITING BLOCK: Breached Cases ────────────────────────────────────────
                # Each line writes one value into the QCL. The letter (B, C, D...) is the
                # column in the QCL Excel sheet. To add a new column here, add a new line
                # following the same pattern and use the correct QCL column letter.
                ws[f"B{current_row}"].value = control_no                                                                             # Col B: Control Check No. (auto-numbered)
                ws[f"C{current_row}"].value = str(case[assignment_group_col]).strip() if is_katowice else format_assignment_group(case[assignment_group_col])  # Col C: Assignment Group
                ws[f"D{current_row}"].value = location                                                                               # Col D: Location
                ws[f"E{current_row}"].value = member_data['name'] if is_katowice else member_data['team_name']                       # Col E: Processor Name
                ws[f"F{current_row}"].value = case_number                                                                            # Col F: Reference Number
                ws[f"G{current_row}"].value = case[hr_service_col]                                                                   # Col G: HR Service
                ws[f"H{current_row}"].value = case[resolution_type_col]                                                              # Col H: SLA Breach Type
                ws[f"J{current_row}"].value = case[breach_reason_col]                                                                # Col J: Breach Reason
                # ── To add a column: ws[f"X{current_row}"].value = case[new_col]
                #    (replace X with the QCL column letter, e.g. "K" for column K)
                # ─────────────────────────────────────────────────────────────────────────

                # Log statistics to terminal, unprocessed notices to GUI console
                if recognized is True:
                    print(f"Breached Cases {control_no}: Case {case_number} from {location} was taken from row {raw_row_num} in {raw_file_name}/Breached Cases")
                elif recognized is False:
                    if log_func:
                        log_func(f"Breached Cases {control_no}: Case {case_number} has an unrecognized location code '{raw_location}'. Taken from row {raw_row_num} in {raw_file_name}/Breached Cases.")
                else:  # recognized is None (blank)
                    if log_func:
                        log_func(f"Breached Cases {control_no}: Case {case_number} has a blank location. Taken from row {raw_row_num} in {raw_file_name}/Breached Cases.")

                control_no += 1
                current_row += 1
                total_cases_written += 1

        if log_func:
            log_func(f"Breached Cases completed: {total_cases_written} cases written.\n")

        return control_no, total_cases_written

    except Exception as e:
        raise Exception(f"Error processing QA Reviews - Breached Cases: {str(e)}")

def process_qa_reviews_csats(qa_review_file, member_file, wb, control_no_start, selected_circle=None, log_func=None, is_katowice=False):

    # Process QA Reviews - CSATs sheet
    # selected_circle: None for all circles, or 1-6 for specific circle
    # log_func: Function to log messages to console
    # Returns: (next_control_no, total_cases_written)

    try:
        # Read the "CSATs" sheet from QA Review file
        csats_df = pd.read_excel(qa_review_file, sheet_name="CSATs")

        # Find required columns in CSATs sheet
        def find_column_containing(df, search_terms):
            """Find column that contains any of the search terms (case-insensitive)"""
            for col in df.columns:
                col_lower = str(col).strip().lower()
                for term in search_terms:
                    # Check if the term is in the column name (handles both bracket and non-bracket formats)
                    if f"[{term.strip().lower()}]" in col_lower or term.strip().lower() in col_lower:
                        return col
            raise ValueError(f"Missing required column containing: {search_terms}")

        # ── COLUMN MAP: CSATs ─────────────────────────────────────────────────────
        # SOURCE KEYWORD (in QA Review file)        QCL OUTPUT COLUMN
        #   "team"                              →  Column C  (Assignment Group)
        #   "subject person country code"       →  Column D  (Location)
        #   "agent"                             →  Column E  (Processor Name)
        #   "csat id"                           →  Column F  (CSAT ID)
        #   "case number"                       →  Column G  (Reference Number)
        #   "hr service"                        →  Column H  (HR Service)
        #   "csatscore" / "csat score"          →  Column I  (CSAT Score)
        #   [CSAT Type]                         →  Column J  (auto-calculated from score)
        #   "commments" (3 M's — not a typo)    →  Column K  (Comment)
        #   [Response Category]                 →  Column M  (auto-calculated from comment)
        #   Note: Columns J and M are auto-calculated — no source column needed.
        #   IMPORTANT: The "commments" column in the raw file has 3 M's ("commments").
        #              This is a known quirk of the source file — do not correct it.
        #
        # TO ADD A COLUMN:
        #   Step 1 — Add a read line below:
        #     new_col = find_column_containing(csats_df, ["exact header keyword"])
        #   Step 2 — Add a write line in the WRITING BLOCK further below:
        #     ws[f"X{current_row}"].value = case[new_col]
        #     (Replace "X" with the target QCL column letter, e.g. "N" for column N)
        #
        # TO REMOVE A COLUMN:
        #   Step 1 — Delete (or comment out) its find_column_containing(...) line below.
        #   Step 2 — Delete (or comment out) its ws[f"X{current_row}"].value = ... line
        #            in the WRITING BLOCK further below.
        #
        # TO RENAME A SOURCE COLUMN (QA Review file header changed):
        #   Update the keyword in find_column_containing(csats_df, ["old name"])
        #   to the new name. Tip: list both as fallbacks: ["new name", "old name"]
        #
        # TO MOVE A QCL COLUMN (column letter changed in QCL template):
        #   In the WRITING BLOCK, change ws[f"OLD_LETTER{current_row}"]
        #   to ws[f"NEW_LETTER{current_row}"]
        # ─────────────────────────────────────────────────────────────────────────
        raw_team_col = find_column_containing(csats_df, ["team"])                          # → QCL Column C
        country_code_col = find_column_containing(csats_df, ["subject person country code"])  # → QCL Column D
        agent_col = find_column_containing(csats_df, ["agent"])                            # → QCL Column E
        csat_id_col = find_column_containing(csats_df, ["csat id"])                        # → QCL Column F
        case_number_col = find_column_containing(csats_df, ["case number"])                # → QCL Column G
        hrservice_col = find_column_containing(csats_df, ["hr service"])                   # → QCL Column H
        csat_score_col = find_column_containing(csats_df, ["csatscore", "csat score"])     # → QCL Column I
        comments_col = find_column_containing(csats_df, ["commments"])  # → QCL Column K  [NOTE: "commments" has 3 M's in the raw file — this is not a typo]

        # Read member list file
        member_header_row = detect_header_row(
            member_file,
            required_columns=["Team", "Tenure"]
        )
        members_df = pd.read_excel(member_file, header=member_header_row)

        # Find formatted name column (with corp key) and Team column (LastName, FirstName format)
        formatted_name_col = None
        for col in members_df.columns:
            sample_values = members_df[col].dropna().astype(str).head(5)
            if any(" - " in val for val in sample_values):
                formatted_name_col = col
                break

        member_name_col = formatted_name_col if formatted_name_col else find_column(members_df, ["Team"])
        team_col = find_column(members_df, ["Team"])  # Column A with "LastName, FirstName" format
        tenure_col = find_column(members_df, ["Tenure"])

        # Build corpkey to member mapping (same as QA Checks)
        corpkey_to_member = {}
        for _, member in members_df.iterrows():
            member_name_raw = str(member[member_name_col]).strip()
            team_name = str(member[team_col]).strip()  # Get the "LastName, FirstName" format
            tenure = member[tenure_col]

            if not member_name_raw or member_name_raw.lower() == 'nan' or pd.isna(tenure):
                continue

            corpkey = extract_corpkey_from_name(member_name_raw)
            if corpkey:
                corpkey_to_member[corpkey] = {
                    'name': member_name_raw,
                    'team_name': team_name,  # Store the Team column name for display
                    'tenure': tenure,
                    'cases': []
                }

        # Get raw file name for logging
        raw_file_name = os.path.basename(qa_review_file)

        # Build dictionary of cases per corpkey with circle filtering (with row index)
        for row_idx, row in csats_df.iterrows():
            agent_name = str(row[agent_col]).strip()
            if agent_name and agent_name.lower() != 'nan':
                # Check if case matches the selected circle/team filter
                if is_katowice:
                    if selected_circle == "All Katowice":
                        include_case = True
                    else:
                        ktw_team = KATOWICE_TEAM_MAP.get(selected_circle, "").lower()
                        include_case = ktw_team in str(row[raw_team_col]).strip().lower()
                else:
                    case_circle = get_circle_from_assignment_group(row[raw_team_col])
                    include_case = selected_circle is None or case_circle == selected_circle
                if include_case:
                    # Extract corpkey from agent field (same as QA Checks with "Assigned to")
                    corpkey = extract_corpkey_from_name(agent_name)
                    if corpkey and corpkey in corpkey_to_member:
                        corpkey_to_member[corpkey]['cases'].append({'row': row, 'raw_row_num': row_idx + 2})

        # Access the CSATs sheet
        if "CSAT" not in wb.sheetnames:
            raise ValueError("QCL template does not contain 'CSAT' sheet")

        ws = wb["CSAT"]
        current_row = find_qcl_start_row(ws, search_column="B", search_text="control check no.")
        control_no = control_no_start

        # Track statistics
        total_cases_written = 0

        print(f"=== PROCESSING CSAT ===")

        # Process each member with cases
        for corpkey, member_data in corpkey_to_member.items():
            cases = member_data['cases']

            if len(cases) == 0:
                continue

            # Write ALL cases to QCL - CSATs sheet (no sampling)
            for case_data in cases:
                case = case_data['row']
                raw_row_num = case_data['raw_row_num']
                case_number = case[case_number_col]
                csat_id = case[csat_id_col]
                location, recognized, raw_location = format_location(case[country_code_col])

                # ── WRITING BLOCK: CSATs ─────────────────────────────────────────────────
                # Each line writes one value into the QCL. The letter (B, C, D...) is the
                # column in the QCL Excel sheet. To add a new column here, add a new line
                # following the same pattern and use the correct QCL column letter.
                ws[f"B{current_row}"].value = control_no                                                                             # Col B: Control Check No. (auto-numbered)
                ws[f"C{current_row}"].value = str(case[raw_team_col]).strip() if is_katowice else format_assignment_group(case[raw_team_col])  # Col C: Assignment Group
                ws[f"D{current_row}"].value = location                                                                               # Col D: Location
                ws[f"E{current_row}"].value = member_data['name'] if is_katowice else member_data['team_name']                       # Col E: Processor Name
                ws[f"F{current_row}"].value = csat_id                                                                                # Col F: CSAT ID
                ws[f"G{current_row}"].value = case_number                                                                            # Col G: Reference Number (HR Case No.)
                ws[f"H{current_row}"].value = case[hrservice_col]                                                                    # Col H: HR Service
                ws[f"I{current_row}"].value = case[csat_score_col]                                                                   # Col I: CSAT Score
                ws[f"J{current_row}"].value = calculate_csat_type(case[csat_score_col])                                              # Col J: CSAT Type (auto-calculated from score)
                ws[f"K{current_row}"].value = case[comments_col]                                                                     # Col K: Comment
                ws[f"M{current_row}"].value = calculate_response_category(case[comments_col])                                        # Col M: Response Category (auto-calculated from comment)
                # ── To add a column: ws[f"X{current_row}"].value = case[new_col]
                #    (replace X with the QCL column letter, e.g. "N" for column N)
                # ─────────────────────────────────────────────────────────────────────────

                # Log statistics to terminal, unprocessed notices to GUI console
                if recognized is True:
                    print(f"CSAT {control_no}: Case {case_number} (CSAT ID: {csat_id}) from {location} was taken from row {raw_row_num} in {raw_file_name}/CSATs")
                elif recognized is False:
                    if log_func:
                        log_func(f"CSAT {control_no}: Case {case_number} (CSAT ID: {csat_id}) has an unrecognized location code '{raw_location}'. Taken from row {raw_row_num} in {raw_file_name}/CSATs.")
                else:  # recognized is None (blank)
                    if log_func:
                        log_func(f"CSAT {control_no}: Case {case_number} (CSAT ID: {csat_id}) has a blank location. Taken from row {raw_row_num} in {raw_file_name}/CSATs.")

                control_no += 1
                current_row += 1
                total_cases_written += 1

        if log_func:
            log_func(f"CSAT completed: {total_cases_written} cases written.\n")

        return control_no, total_cases_written

    except Exception as e:
        raise Exception(f"Error processing QA Reviews - CSATs: {str(e)}")

def process_qa_reviews_csat_chat(qa_review_file, member_file, wb, control_no_start, selected_circle_text="All Circles", log_func=None, is_katowice=False):

    # Process QA Reviews - CSAT Chat sheet
    # Only processes if All Circles or Circle 6 (Contact Center) is selected
    # log_func: Function to log messages to console
    # Returns: (next_control_no, total_cases_written)

    # CSAT Chat is exclusive to Circle 6 Contact Center - skip if another circle is selected
    if selected_circle_text != "All Circles" and selected_circle_text != "Circle 6 (Contact Center)":
        print("CSAT Chat: Skipped (only applicable for Circle 6 Contact Center).")
        return control_no_start, 0

    try:
        # Check if "CSAT Chat" sheet exists in the QA Review file
        xls = pd.ExcelFile(qa_review_file)
        if "CSAT Chat" not in xls.sheet_names:
            if log_func:
                log_func("CSAT Chat: No 'CSAT Chat' sheet found in the QA Reviews file.")
            return control_no_start, 0

        # Read the "CSAT Chat" sheet from QA Review file
        csat_chat_df = pd.read_excel(qa_review_file, sheet_name="CSAT Chat")

        # Check if the sheet has any data
        if csat_chat_df.empty:
            if log_func:
                log_func("CSAT Chat: No data found in the 'CSAT Chat' sheet.")
            return control_no_start, 0

        # Find required columns in CSAT Chat sheet
        def find_column_containing(df, search_terms):
            """Find column that contains any of the search terms (case-insensitive)"""
            for col in df.columns:
                col_lower = str(col).strip().lower()
                for term in search_terms:
                    if f"[{term.strip().lower()}]" in col_lower or term.strip().lower() in col_lower:
                        return col
            raise ValueError(f"Missing required column containing: {search_terms}")

        # ── COLUMN MAP: CSAT Chat ─────────────────────────────────────────────────
        # This sheet is exclusive to Circle 6 (Contact Center).
        # SOURCE KEYWORD (in QA Review file)        QCL OUTPUT COLUMN
        #   "team"                              →  Column C  (Assignment Group)
        #   "subject person country code"       →  Column D  (Location)
        #   "agent"                             →  Column E  (Processor Name)
        #   "csat id"                           →  Column F  (CSAT ID)
        #   "case number"                       →  Column G  (Reference Number)
        #   "hr service"                        →  Column H  (HR Service)
        #   "csatscore" / "csat score"          →  Column I  (CSAT Score)
        #   [CSAT Type]                         →  Column J  (auto-calculated from score)
        #   "commments" (3 M's — not a typo)    →  Column K  (Comment)
        #   [Response Category]                 →  Column M  (auto-calculated from comment)
        #   Note: Columns J and M are auto-calculated — no source column needed.
        #   IMPORTANT: The "commments" column in the raw file has 3 M's ("commments").
        #              This is a known quirk of the source file — do not correct it.
        #
        # TO ADD A COLUMN:
        #   Step 1 — Add a read line below:
        #     new_col = find_column_containing(csat_chat_df, ["exact header keyword"])
        #   Step 2 — Add a write line in the WRITING BLOCK further below:
        #     ws[f"X{current_row}"].value = case[new_col]
        #     (Replace "X" with the target QCL column letter, e.g. "N" for column N)
        #
        # TO REMOVE A COLUMN:
        #   Step 1 — Delete (or comment out) its find_column_containing(...) line below.
        #   Step 2 — Delete (or comment out) its ws[f"X{current_row}"].value = ... line
        #            in the WRITING BLOCK further below.
        #
        # TO RENAME A SOURCE COLUMN (QA Review file header changed):
        #   Update the keyword in find_column_containing(csat_chat_df, ["old name"])
        #   to the new name. Tip: list both as fallbacks: ["new name", "old name"]
        #
        # TO MOVE A QCL COLUMN (column letter changed in QCL template):
        #   In the WRITING BLOCK, change ws[f"OLD_LETTER{current_row}"]
        #   to ws[f"NEW_LETTER{current_row}"]
        # ─────────────────────────────────────────────────────────────────────────
        raw_team_col = find_column_containing(csat_chat_df, ["team"])                          # → QCL Column C
        country_code_col = find_column_containing(csat_chat_df, ["subject person country code"])  # → QCL Column D
        agent_col = find_column_containing(csat_chat_df, ["agent"])                            # → QCL Column E
        csat_id_col = find_column_containing(csat_chat_df, ["csat id"])                        # → QCL Column F
        case_number_col = find_column_containing(csat_chat_df, ["case number"])                # → QCL Column G
        hrservice_col = find_column_containing(csat_chat_df, ["hr service"])                   # → QCL Column H
        csat_score_col = find_column_containing(csat_chat_df, ["csatscore", "csat score"])     # → QCL Column I
        comments_col = find_column_containing(csat_chat_df, ["commments"])  # → QCL Column K  [NOTE: "commments" has 3 M's in the raw file — this is not a typo]

        # Read member list file
        member_header_row = detect_header_row(
            member_file,
            required_columns=["Team", "Tenure"]
        )
        members_df = pd.read_excel(member_file, header=member_header_row)

        # Find formatted name column (with corp key) and Team column (LastName, FirstName format)
        formatted_name_col = None
        for col in members_df.columns:
            sample_values = members_df[col].dropna().astype(str).head(5)
            if any(" - " in val for val in sample_values):
                formatted_name_col = col
                break

        member_name_col = formatted_name_col if formatted_name_col else find_column(members_df, ["Team"])
        team_col = find_column(members_df, ["Team"])
        tenure_col = find_column(members_df, ["Tenure"])

        # Build corpkey to member mapping
        corpkey_to_member = {}
        for _, member in members_df.iterrows():
            member_name_raw = str(member[member_name_col]).strip()
            team_name = str(member[team_col]).strip()
            tenure = member[tenure_col]

            if not member_name_raw or member_name_raw.lower() == 'nan' or pd.isna(tenure):
                continue

            corpkey = extract_corpkey_from_name(member_name_raw)
            if corpkey:
                corpkey_to_member[corpkey] = {
                    'name': member_name_raw,
                    'team_name': team_name,
                    'tenure': tenure,
                    'cases': []
                }

        # Get raw file name for logging
        raw_file_name = os.path.basename(qa_review_file)

        # Build dictionary of cases per corpkey (with row index)
        # No circle filtering on individual cases — the entire sheet is Circle 6 only
        for row_idx, row in csat_chat_df.iterrows():
            agent_name = str(row[agent_col]).strip()
            if agent_name and agent_name.lower() != 'nan':
                corpkey = extract_corpkey_from_name(agent_name)
                if corpkey and corpkey in corpkey_to_member:
                    corpkey_to_member[corpkey]['cases'].append({'row': row, 'raw_row_num': row_idx + 2})

        # Access the CSAT Chat sheet in QCL
        if "CSAT Chat" not in wb.sheetnames:
            raise ValueError("QCL template does not contain 'CSAT Chat' sheet")

        ws = wb["CSAT Chat"]
        current_row = find_qcl_start_row(ws, search_column="B", search_text="control check no.")
        control_no = control_no_start

        # Track statistics
        total_cases_written = 0

        print(f"=== PROCESSING CSAT CHAT ===")

        # Process each member with cases
        for corpkey, member_data in corpkey_to_member.items():
            cases = member_data['cases']

            if len(cases) == 0:
                continue

            # Write ALL cases to QCL - CSAT Chat sheet (no sampling)
            for case_data in cases:
                case = case_data['row']
                raw_row_num = case_data['raw_row_num']
                case_number = case[case_number_col]
                csat_id = case[csat_id_col]
                location, recognized, raw_location = format_location(case[country_code_col])

                # ── WRITING BLOCK: CSAT Chat ─────────────────────────────────────────────
                # Each line writes one value into the QCL. The letter (B, C, D...) is the
                # column in the QCL Excel sheet. To add a new column here, add a new line
                # following the same pattern and use the correct QCL column letter.
                ws[f"B{current_row}"].value = control_no                                                                             # Col B: Control Check No. (auto-numbered)
                ws[f"C{current_row}"].value = str(case[raw_team_col]).strip() if is_katowice else format_assignment_group(case[raw_team_col])  # Col C: Assignment Group
                ws[f"D{current_row}"].value = location                                                                               # Col D: Location
                ws[f"E{current_row}"].value = member_data['name'] if is_katowice else member_data['team_name']                       # Col E: Processor Name
                ws[f"F{current_row}"].value = csat_id                                                                                # Col F: CSAT ID
                ws[f"G{current_row}"].value = case_number                                                                            # Col G: Reference Number (HR Case No.)
                ws[f"H{current_row}"].value = case[hrservice_col]                                                                    # Col H: HR Service
                ws[f"I{current_row}"].value = case[csat_score_col]                                                                   # Col I: CSAT Score
                ws[f"J{current_row}"].value = calculate_csat_type(case[csat_score_col])                                              # Col J: CSAT Type (auto-calculated from score)
                ws[f"K{current_row}"].value = case[comments_col]                                                                     # Col K: Comment
                ws[f"M{current_row}"].value = calculate_response_category(case[comments_col])                                        # Col M: Response Category (auto-calculated from comment)
                # ── To add a column: ws[f"X{current_row}"].value = case[new_col]
                #    (replace X with the QCL column letter, e.g. "N" for column N)
                # ─────────────────────────────────────────────────────────────────────────

                # Log statistics to terminal, unprocessed notices to GUI console
                if recognized is True:
                    print(f"CSAT Chat {control_no}: Case {case_number} (CSAT ID: {csat_id}) from {location} was taken from row {raw_row_num} in {raw_file_name}/CSAT Chat")
                elif recognized is False:
                    if log_func:
                        log_func(f"CSAT Chat {control_no}: Case {case_number} (CSAT ID: {csat_id}) has an unrecognized location code '{raw_location}'. Taken from row {raw_row_num} in {raw_file_name}/CSAT Chat.")
                else:  # recognized is None (blank)
                    if log_func:
                        log_func(f"CSAT Chat {control_no}: Case {case_number} (CSAT ID: {csat_id}) has a blank location. Taken from row {raw_row_num} in {raw_file_name}/CSAT Chat.")

                control_no += 1
                current_row += 1
                total_cases_written += 1

        if log_func:
            log_func(f"CSAT Chat completed: {total_cases_written} cases written.\n")

        return control_no, total_cases_written

    except Exception as e:
        raise Exception(f"Error processing QA Reviews - CSAT Chat: {str(e)}")

def process_qa_reviews_live_chat_fcr(qa_review_file, member_file, wb, control_no_start, selected_circle_text="All Circles", log_func=None, is_katowice=False):

    # Process QA Reviews - Live Chat FCR sheet
    # Only processes if All Circles or Circle 6 (Contact Center) is selected
    # log_func: Function to log messages to console
    # Returns: (next_control_no, total_cases_written)

    # Live Chat FCR is exclusive to Circle 6 Contact Center - skip if another circle is selected
    if selected_circle_text != "All Circles" and selected_circle_text != "Circle 6 (Contact Center)":
        print("Live Chat FCR: Skipped (only applicable for Circle 6 Contact Center).")
        return control_no_start, 0

    try:
        # Check if "Live Chat FCR" sheet exists in the QA Review file
        xls = pd.ExcelFile(qa_review_file)
        if "Live Chat FCR" not in xls.sheet_names:
            if log_func:
                log_func("Live Chat FCR: Sheet not found in QA Reviews file. Skipping.")
            return control_no_start, 0

        # Detect header row and read the Live Chat FCR sheet
        header_row = detect_header_row(
            qa_review_file,
            sheet_name="Live Chat FCR",
            required_columns=["Team", "Agent", "Case Number"]
        )
        live_chat_fcr_df = pd.read_excel(qa_review_file, sheet_name="Live Chat FCR", header=header_row)

        # Check if sheet has data
        if live_chat_fcr_df.empty:
            if log_func:
                log_func("Live Chat FCR: Sheet is empty. Skipping.")
            return control_no_start, 0

        # Helper function for bracket-aware column search
        def find_column_containing(df, search_terms):
            for col in df.columns:
                col_check = str(col).strip().lower()
                # Check for bracket format like "F09 - SNOW HR Cases[Column Name]"
                if "[" in col_check and "]" in col_check:
                    bracket_content = col_check[col_check.index("[") + 1:col_check.index("]")]
                    for term in search_terms:
                        if term.lower() == bracket_content:
                            return col
                for term in search_terms:
                    if term.lower() == col_check:
                        return col
            raise ValueError(f"Could not find column matching any of {search_terms} in Live Chat FCR sheet")

        # ── COLUMN MAP: Live Chat FCR ─────────────────────────────────────────────
        # This sheet is exclusive to Circle 6 (Contact Center).
        # SOURCE KEYWORD (in QA Review file)        QCL OUTPUT COLUMN
        #   "team"                              →  Column C  (Assignment Group)
        #   "subject person country code"       →  Column D  (Location / Country)
        #   "agent"                             →  Column E  (Processor Name)
        #   "csat id"                           →  Column F  (CSAT ID)
        #   "case number"                       →  Column G  (Reference Number)
        #   "hr service"                        →  Column H  (HR Service)
        #   "commments" (3 M's — not a typo)    →  Column I  (Comment)
        #   "fcr"                               →  Column J  (FCR)
        #   IMPORTANT: The "commments" column in the raw file has 3 M's ("commments").
        #              This is a known quirk of the source file — do not correct it.
        #
        # TO ADD A COLUMN:
        #   Step 1 — Add a read line below:
        #     new_col = find_column_containing(live_chat_fcr_df, ["exact header keyword"])
        #   Step 2 — Add a write line in the WRITING BLOCK further below:
        #     ws[f"X{current_row}"].value = case[new_col]
        #     (Replace "X" with the target QCL column letter, e.g. "K" for column K)
        #
        # TO REMOVE A COLUMN:
        #   Step 1 — Delete (or comment out) its find_column_containing(...) line below.
        #   Step 2 — Delete (or comment out) its ws[f"X{current_row}"].value = ... line
        #            in the WRITING BLOCK further below.
        #
        # TO RENAME A SOURCE COLUMN (QA Review file header changed):
        #   Update the keyword in find_column_containing(live_chat_fcr_df, ["old name"])
        #   to the new name. Tip: list both as fallbacks: ["new name", "old name"]
        #
        # TO MOVE A QCL COLUMN (column letter changed in QCL template):
        #   In the WRITING BLOCK, change ws[f"OLD_LETTER{current_row}"]
        #   to ws[f"NEW_LETTER{current_row}"]
        # ─────────────────────────────────────────────────────────────────────────
        raw_team_col = find_column_containing(live_chat_fcr_df, ["team"])                          # → QCL Column C
        country_code_col = find_column_containing(live_chat_fcr_df, ["subject person country code"])  # → QCL Column D
        agent_col = find_column_containing(live_chat_fcr_df, ["agent"])                            # → QCL Column E
        csat_id_col = find_column_containing(live_chat_fcr_df, ["csat id"])                        # → QCL Column F
        case_number_col = find_column_containing(live_chat_fcr_df, ["case number"])                # → QCL Column G
        hrservice_col = find_column_containing(live_chat_fcr_df, ["hr service"])                   # → QCL Column H
        comments_col = find_column_containing(live_chat_fcr_df, ["commments"])  # → QCL Column I  [NOTE: "commments" has 3 M's in the raw file — this is not a typo]
        fcr_col = find_column_containing(live_chat_fcr_df, ["fcr"])                                # → QCL Column J

        # Read member list file
        member_header_row = detect_header_row(
            member_file,
            required_columns=["Team", "Tenure"]
        )
        members_df = pd.read_excel(member_file, header=member_header_row)

        # Find formatted name column (with corp key) and Team column (LastName, FirstName format)
        formatted_name_col = None
        for col in members_df.columns:
            sample_values = members_df[col].dropna().astype(str).head(5)
            if any(" - " in val for val in sample_values):
                formatted_name_col = col
                break

        member_name_col = formatted_name_col if formatted_name_col else find_column(members_df, ["Team"])
        team_col = find_column(members_df, ["Team"])
        tenure_col = find_column(members_df, ["Tenure"])

        # Build corpkey to member mapping
        corpkey_to_member = {}
        for _, member in members_df.iterrows():
            member_name_raw = str(member[member_name_col]).strip()
            team_name = str(member[team_col]).strip()
            tenure = member[tenure_col]

            if not member_name_raw or member_name_raw.lower() == 'nan' or pd.isna(tenure):
                continue

            corpkey = extract_corpkey_from_name(member_name_raw)
            if corpkey:
                corpkey_to_member[corpkey] = {
                    'name': member_name_raw,
                    'team_name': team_name,
                    'tenure': tenure,
                    'cases': []
                }

        # Get raw file name for logging
        raw_file_name = os.path.basename(qa_review_file)

        # Build dictionary of cases per corpkey (no per-case circle filtering — entire sheet is Circle 6 Contact Center)
        for row_idx, row in live_chat_fcr_df.iterrows():
            agent_name = str(row[agent_col]).strip()
            if agent_name and agent_name.lower() != 'nan':
                corpkey = extract_corpkey_from_name(agent_name)
                if corpkey and corpkey in corpkey_to_member:
                    corpkey_to_member[corpkey]['cases'].append({'row': row, 'raw_row_num': row_idx + 2})

        # Access the Live Chat FCR sheet in QCL
        if "Live Chat FCR" not in wb.sheetnames:
            raise ValueError("QCL template does not contain 'Live Chat FCR' sheet")

        ws = wb["Live Chat FCR"]
        current_row = find_qcl_start_row(ws, search_column="B", search_text="control check no.")
        control_no = control_no_start

        # Track statistics
        total_cases_written = 0

        print(f"=== PROCESSING LIVE CHAT FCR ===")

        # Process each member with cases
        for corpkey, member_data in corpkey_to_member.items():
            cases = member_data['cases']

            if len(cases) == 0:
                continue

            # Write ALL cases to QCL - Live Chat FCR sheet (no sampling)
            for case_data in cases:
                case = case_data['row']
                raw_row_num = case_data['raw_row_num']
                case_number = case[case_number_col]
                csat_id = case[csat_id_col]
                location, recognized, raw_location = format_location(case[country_code_col])

                # ── WRITING BLOCK: Live Chat FCR ─────────────────────────────────────────
                # Each line writes one value into the QCL. The letter (B, C, D...) is the
                # column in the QCL Excel sheet. To add a new column here, add a new line
                # following the same pattern and use the correct QCL column letter.
                ws[f"B{current_row}"].value = control_no                                                                             # Col B: Control Check No. (auto-numbered)
                ws[f"C{current_row}"].value = str(case[raw_team_col]).strip() if is_katowice else format_assignment_group(case[raw_team_col])  # Col C: Assignment Group
                ws[f"D{current_row}"].value = location                                                                               # Col D: Location / Country
                ws[f"E{current_row}"].value = member_data['name'] if is_katowice else member_data['team_name']                       # Col E: Processor Name
                ws[f"F{current_row}"].value = csat_id                                                                                # Col F: CSAT ID
                ws[f"G{current_row}"].value = case_number                                                                            # Col G: Reference Number (HR Case No.)
                ws[f"H{current_row}"].value = case[hrservice_col]                                                                    # Col H: HR Service
                ws[f"I{current_row}"].value = case[comments_col]                                                                     # Col I: Comment
                ws[f"J{current_row}"].value = case[fcr_col]                                                                          # Col J: FCR
                # ── To add a column: ws[f"X{current_row}"].value = case[new_col]
                #    (replace X with the QCL column letter, e.g. "K" for column K)
                # ─────────────────────────────────────────────────────────────────────────

                # Log statistics to terminal, unprocessed notices to GUI console
                if recognized is True:
                    print(f"Live Chat FCR {control_no}: Case {case_number} (CSAT ID: {csat_id}) from {location} was taken from row {raw_row_num} in {raw_file_name}/Live Chat FCR")
                elif recognized is False:
                    if log_func:
                        log_func(f"Live Chat FCR {control_no}: Case {case_number} (CSAT ID: {csat_id}) has an unrecognized location code '{raw_location}'. Taken from row {raw_row_num} in {raw_file_name}/Live Chat FCR.")
                else:  # recognized is None (blank)
                    if log_func:
                        log_func(f"Live Chat FCR {control_no}: Case {case_number} (CSAT ID: {csat_id}) has a blank location. Taken from row {raw_row_num} in {raw_file_name}/Live Chat FCR.")

                control_no += 1
                current_row += 1
                total_cases_written += 1

        if log_func:
            log_func(f"Live Chat FCR completed: {total_cases_written} cases written.\n")

        return control_no, total_cases_written

    except Exception as e:
        raise Exception(f"Error processing QA Reviews - Live Chat FCR: {str(e)}")

def process_qa_reviews_wd_breach(qa_review_file, wb, control_no_start, selected_circle_text="All Katowice", log_func=None):

    # Process QA Reviews - WD Breach sheet
    # Only processes when a Katowice option is selected
    # No processor name — Column F is left blank for manual input
    # Returns: (next_control_no, total_cases_written)

    # WD Breach is exclusive to Katowice — skip if a Manila circle is selected
    if not is_katowice_selection(selected_circle_text):
        print("WD Breach: Skipped (only applicable for Katowice).")
        return control_no_start, 0

    try:
        # Check if "WD Breach" sheet exists in the QA Review file
        xls = pd.ExcelFile(qa_review_file)
        if "WD Breach" not in xls.sheet_names:
            if log_func:
                log_func("WD Breach: Sheet not found in QA Reviews file. Skipping.")
            return control_no_start, 0

        # Detect header row — WD Breach uses "F02 - Workday Business Process[ColumnName]" format
        # Search for a row containing any cell with "[Country]" in it
        preview = pd.read_excel(qa_review_file, sheet_name="WD Breach", header=None)
        header_row = None
        for idx, row in preview.iterrows():
            row_values = [str(cell).strip().lower() for cell in row.values if pd.notna(cell)]
            if any("[country]" in val for val in row_values):
                header_row = idx
                break
        if header_row is None:
            raise ValueError("Could not find header row in WD Breach sheet")
        wd_breach_df = pd.read_excel(qa_review_file, sheet_name="WD Breach", header=header_row)

        # Check if sheet has data
        if wd_breach_df.empty:
            if log_func:
                log_func("WD Breach: Sheet is empty. Skipping.")
            return control_no_start, 0

        # Helper function for bracket-aware column search
        def find_column_containing(df, search_terms):
            for col in df.columns:
                col_check = str(col).strip().lower()
                if "[" in col_check and "]" in col_check:
                    bracket_content = col_check[col_check.index("[") + 1:col_check.index("]")]
                    for term in search_terms:
                        if term.lower() == bracket_content:
                            return col
                for term in search_terms:
                    if term.lower() == col_check:
                        return col
            raise ValueError(f"Could not find column matching any of {search_terms} in WD Breach sheet")

        # ── COLUMN MAP: WD Breached Cases ────────────────────────────────────────
        # This sheet is exclusive to Katowice.
        # Unlike other sheets, there is no member-matching — Column F (Processor Name)
        # is intentionally left blank for manual input after export.
        # SOURCE KEYWORD (in QA Review file)        QCL OUTPUT COLUMN
        #   "snow assignment group"             →  Column C  (Team)
        #   "overall functional area"           →  Column D  (Functional Area)
        #   "country"                           →  Column E  (Country)
        #   [Column F intentionally left blank] →  Column F  (Processor Name — fill manually)
        #   "business process name"             →  Column G  (Business Process Name)
        #   "workday id"                        →  Column H  (Reference Number / Workday ID)
        #   "bp completed date"                 →  Column I  (BP Completed Date)
        #   "target sla days"                   →  Column J  (Target SLA Days)
        #
        # TO ADD A COLUMN:
        #   Step 1 — Add a read line below:
        #     new_col = find_column_containing(wd_breach_df, ["exact header keyword"])
        #   Step 2 — Add a write line in the WRITING BLOCK further below:
        #     ws[f"X{current_row}"].value = case[new_col]
        #     (Replace "X" with the target QCL column letter, e.g. "K" for column K)
        #
        # TO REMOVE A COLUMN:
        #   Step 1 — Delete (or comment out) its find_column_containing(...) line below.
        #   Step 2 — Delete (or comment out) its ws[f"X{current_row}"].value = ... line
        #            in the WRITING BLOCK further below.
        #
        # TO RENAME A SOURCE COLUMN (QA Review file header changed):
        #   Update the keyword in find_column_containing(wd_breach_df, ["old name"])
        #   to the new name. Tip: list both as fallbacks: ["new name", "old name"]
        #
        # TO MOVE A QCL COLUMN (column letter changed in QCL template):
        #   In the WRITING BLOCK, change ws[f"OLD_LETTER{current_row}"]
        #   to ws[f"NEW_LETTER{current_row}"]
        # ─────────────────────────────────────────────────────────────────────────
        snow_assignment_col = find_column_containing(wd_breach_df, ["snow assignment group"])  # → QCL Column C
        functional_area_col = find_column_containing(wd_breach_df, ["overall functional area"])  # → QCL Column D
        country_col = find_column_containing(wd_breach_df, ["country"])                        # → QCL Column E
        bp_name_col = find_column_containing(wd_breach_df, ["business process name"])          # → QCL Column G
        workday_id_col = find_column_containing(wd_breach_df, ["workday id"])                  # → QCL Column H
        bp_completed_col = find_column_containing(wd_breach_df, ["bp completed date"])         # → QCL Column I
        target_sla_col = find_column_containing(wd_breach_df, ["target sla days"])             # → QCL Column J

        # Get raw file name for logging
        raw_file_name = os.path.basename(qa_review_file)

        # Filter cases by Katowice team if a specific team is selected
        selected_circle = parse_circle_selection(selected_circle_text)
        if selected_circle_text != "All Katowice" and selected_circle in KATOWICE_TEAM_MAP:
            katowice_team = KATOWICE_TEAM_MAP[selected_circle].lower()
            wd_breach_df = wd_breach_df[
                wd_breach_df[snow_assignment_col].astype(str).str.strip().str.lower().str.contains(katowice_team, na=False)
            ]

        # Access the WD Breached Cases sheet in QCL
        if "WD Breached Cases" not in wb.sheetnames:
            raise ValueError("QCL template does not contain 'WD Breached Cases' sheet")

        ws = wb["WD Breached Cases"]
        current_row = find_qcl_start_row(ws, search_column="B", search_text="control check no.")
        control_no = control_no_start

        # Track statistics
        total_cases_written = 0

        print(f"=== PROCESSING WD BREACH ===")

        # Process each case (no member matching — processor name is manual)
        for row_idx, case in wd_breach_df.iterrows():
            raw_row_num = row_idx + 2
            workday_id = case[workday_id_col]
            country = case[country_col] if not pd.isna(case[country_col]) else ""

            # ── WRITING BLOCK: WD Breached Cases ─────────────────────────────────────
            # Each line writes one value into the QCL. The letter (B, C, D...) is the
            # column in the QCL Excel sheet. To add a new column here, add a new line
            # following the same pattern and use the correct QCL column letter.
            ws[f"B{current_row}"].value = control_no                                                                                  # Col B: Control Check No. (auto-numbered)
            ws[f"C{current_row}"].value = str(case[snow_assignment_col]).strip()                                                      # Col C: Team (no formatting — Katowice only)
            ws[f"D{current_row}"].value = str(case[functional_area_col]).strip() if not pd.isna(case[functional_area_col]) else ""    # Col D: Functional Area
            ws[f"E{current_row}"].value = str(country).strip()                                                                        # Col E: Country
            # Col F: Processor Name — intentionally left blank for manual input after export
            ws[f"G{current_row}"].value = str(case[bp_name_col]).strip() if not pd.isna(case[bp_name_col]) else ""                   # Col G: Business Process Name
            ws[f"H{current_row}"].value = workday_id                                                                                  # Col H: Reference Number (Workday ID)
            ws[f"I{current_row}"].value = case[bp_completed_col]                                                                      # Col I: BP Completed Date
            ws[f"J{current_row}"].value = case[target_sla_col]                                                                        # Col J: Target SLA Days
            # ── To add a column: ws[f"X{current_row}"].value = case[new_col]
            #    (replace X with the QCL column letter, e.g. "K" for column K)
            # ─────────────────────────────────────────────────────────────────────────

            print(f"WD Breach {control_no}: Workday ID {workday_id} from {country} was taken from row {raw_row_num} in {raw_file_name}/WD Breach")

            control_no += 1
            current_row += 1
            total_cases_written += 1

        if log_func:
            log_func(f"WD Breach completed: {total_cases_written} cases written.\n")

        return control_no, total_cases_written

    except Exception as e:
        raise Exception(f"Error processing QA Reviews - WD Breach: {str(e)}")

def process_files():
    raw_file = qa_checks_entry.get()
    qa_review_file = qa_review_entry.get()
    qcl_template = qcl_entry.get()
    member_file = member_entry.get()
    selected_circle_text = circle_var.get()

    # Require Member List and QCL Template
    if not member_file or not qcl_template:
        messagebox.showerror("Error", "Please select Member List and QCL Template files.")
        return

    # Require at least one raw data file (QA Checks or QA Reviews)
    if not raw_file and not qa_review_file:
        messagebox.showerror("Error", "Please select at least one raw data file (QA Checks or QA Reviews).")
        return

    # Parse the selected circle
    selected_circle = parse_circle_selection(selected_circle_text)
    is_katowice = is_katowice_selection(selected_circle_text)

    # Build filter display message
    if is_katowice:
        circle_filter_msg = selected_circle_text
    elif selected_circle:
        circle_filter_msg = f"Circle {selected_circle}"
    else:
        circle_filter_msg = "All Circles"

    # Validate Katowice selection against member list
    if is_katowice and member_file:
        try:
            member_header_row = detect_header_row(member_file, required_columns=["Team"])
            member_preview = pd.read_excel(member_file, header=member_header_row, nrows=0)
            has_ktw = any("ktw support" in str(col).strip().lower() for col in member_preview.columns)
            if not has_ktw:
                messagebox.showerror("Error", "Katowice option selected but the Member List does not contain a 'KTW Support' column.")
                return
        except Exception as e:
            messagebox.showerror("Error", f"Could not read Member List to validate Katowice support: {str(e)}")
            return

    # Clear console and start logging
    clear_console()
    print(f"\n========== PROCESSING WITH FILTER: {circle_filter_msg} ==========")

    try:
        # Copy template and preserve all features first
        file_ext = os.path.splitext(qcl_template)[1]
        output_path = qcl_template.replace(file_ext, f"_POPULATED{file_ext}")
        shutil.copy2(qcl_template, output_path)

        wb = load_workbook(
            output_path,
            keep_vba=True,
            data_only=False,
            keep_links=True
        )

        # Initialize statistics
        qa_checks_stats = None
        qa_reviews_reopened_stats = None
        qa_reviews_breached_stats = None
        qa_reviews_csats_stats = None
        qa_reviews_csat_chat_stats = None
        qa_reviews_live_chat_fcr_stats = None
        qa_reviews_wd_breach_stats = None

        # ========== PROCESS QA CHECKS ==========
        if raw_file:
            print("\n========== PROCESSING QA CHECKS ==========")

            # Read raw data file
            raw_df = pd.read_excel(raw_file)

            # ── COLUMN MAP: QA Checks ─────────────────────────────────────────────────
            # NOTE: QA Checks uses EXACT column name matching (find_column), unlike the
            # QA Review sheets which use partial/keyword matching (find_column_containing).
            # The column name in the source file must match exactly (case-insensitive).
            #
            # SOURCE COLUMN NAME (in QA Checks file)    QCL OUTPUT COLUMN
            #   "Location"                          →  Column D  (Location)
            #   "Assigned to" / "Assigned To"       →  Column E  (Processor Name)
            #   "Number"                            →  Column F  (Reference Number)
            #   "HR Service"                        →  Column G  (HR Service)
            #   "Assignment group"                  →  Column C  (Assignment Group, used for circle filtering)
            #   "Opened for" / "Opened For"         →  (fallback location lookup — not written directly to QCL)
            #   Note: Column B is the auto-numbered Control Check No.
            #
            # TO ADD A COLUMN:
            #   Step 1 — Add a read line below:
            #     new_col = find_column(raw_df, ["Exact Column Header Name"])
            #   Step 2 — Add a write line in the WRITING BLOCK further below:
            #     ws[f"X{current_row}"].value = case[new_col]
            #     (Replace "X" with the target QCL column letter, e.g. "H" for column H)
            #
            # TO REMOVE A COLUMN:
            #   Step 1 — Delete (or comment out) its find_column(...) line below.
            #   Step 2 — Delete (or comment out) its ws[f"X{current_row}"].value = ... line
            #            in the WRITING BLOCK further below.
            #
            # TO RENAME A SOURCE COLUMN (QA Checks file header changed):
            #   Update the name inside find_column(raw_df, ["old name"])
            #   to the new name. Tip: list both as fallbacks: ["new name", "old name"]
            #
            # TO MOVE A QCL COLUMN (column letter changed in QCL template):
            #   In the WRITING BLOCK, change ws[f"OLD_LETTER{current_row}"]
            #   to ws[f"NEW_LETTER{current_row}"]
            # ─────────────────────────────────────────────────────────────────────────
            location_col = find_column(raw_df, ["Location"])                         # → used for QCL Column D (via format_location)
            assigned_col = find_column(raw_df, ["Assigned to", "Assigned To"])       # → QCL Column E (Processor Name)
            case_id_col = find_column(raw_df, ["Number"])                            # → QCL Column F (Reference Number)
            service_col = find_column(raw_df, ["HR Service"])                        # → QCL Column G (HR Service)
            assignment_group_col = find_column(raw_df, ["Assignment group"])         # → QCL Column C (Assignment Group, also used for circle filtering)
            opened_for_col = find_column(raw_df, ["Opened for", "Opened For"])       # → fallback location lookup only (not written directly to QCL)

            # Read member list file
            member_header_row = detect_header_row(
                member_file,
                required_columns=["Team"]
            )
            members_df = pd.read_excel(member_file, header=member_header_row)

            # Find formatted name column (with corp key) and Team column (LastName, FirstName format)
            formatted_name_col = None
            for col in members_df.columns:
                sample_values = members_df[col].dropna().astype(str).head(5)
                if any(" - " in val for val in sample_values):
                    formatted_name_col = col
                    break

            member_name_col = formatted_name_col if formatted_name_col else find_column(members_df, ["Team"])
            team_col = find_column(members_df, ["Team"])  # Column A with "LastName, FirstName" format

            # Build corpkey to member mapping
            corpkey_to_member = {}
            for _, member in members_df.iterrows():
                member_name_raw = str(member[member_name_col]).strip()
                team_name = str(member[team_col]).strip()  # Get the "LastName, FirstName" format

                if not member_name_raw or member_name_raw.lower() == 'nan':
                    continue

                corpkey = extract_corpkey_from_name(member_name_raw)
                if corpkey:
                    corpkey_to_member[corpkey] = {
                        'name': member_name_raw,
                        'team_name': team_name,  # Store the Team column name for display
                        'cases': []
                    }

            # Get raw file name for logging
            raw_file_name = os.path.basename(raw_file)

            # Build dictionary of cases per corpkey (with row index for logging)
            for row_idx, row in raw_df.iterrows():
                assigned_to = str(row[assigned_col]).strip()
                if assigned_to and assigned_to.lower() != 'nan':
                    # Check if case matches the selected circle/team filter
                    if is_katowice:
                        if selected_circle == "All Katowice":
                            include_case = True
                        else:
                            ktw_team = KATOWICE_TEAM_MAP.get(selected_circle, "").lower()
                            include_case = ktw_team in str(row[assignment_group_col]).strip().lower()
                    else:
                        case_circle = get_circle_from_assignment_group(row[assignment_group_col])
                        include_case = selected_circle is None or case_circle == selected_circle
                    if include_case:
                        # Extract corpkey from assigned_to field
                        corpkey = extract_corpkey_from_name(assigned_to)
                        if corpkey and corpkey in corpkey_to_member:
                            # Store row with its index (add 2 for Excel row: 1 for header, 1 for 0-index)
                            corpkey_to_member[corpkey]['cases'].append({'row': row, 'raw_row_num': row_idx + 2})

            print(f"=== PROCESSING QA CHECKS ===")

            ws = wb["QA Checks"]
            current_row = find_qcl_start_row(ws)
            control_no = 1

            # Track processing statistics
            total_cases_written = 0

            # Process each member with cases
            for corpkey, member_data in corpkey_to_member.items():
                cases = member_data['cases']

                if len(cases) == 0:
                    continue

                # Fixed cap of 15 cases per agent; randomize only if agent has more than 15
                if len(cases) > 15:
                    sampled_cases = random.sample(cases, 15)
                else:
                    sampled_cases = cases

                # Write sampled cases to QCL
                for case_data in sampled_cases:
                    case = case_data['row']
                    raw_row_num = case_data['raw_row_num']
                    case_number = case[case_id_col]
                    location, recognized, raw_location = get_case_location(case[location_col], case[opened_for_col])

                    # ── WRITING BLOCK: QA Checks ─────────────────────────────────────────────
                    # Each line writes one value into the QCL. The letter (B, C, D...) is the
                    # column in the QCL Excel sheet. To add a new column here, add a new line
                    # following the same pattern and use the correct QCL column letter.
                    ws[f"B{current_row}"].value = control_no                                                                             # Col B: Control Check No. (auto-numbered)
                    ws[f"C{current_row}"].value = str(case[assignment_group_col]).strip() if is_katowice else format_assignment_group(case[assignment_group_col])  # Col C: Assignment Group
                    ws[f"D{current_row}"].value = location                                                                               # Col D: Location
                    ws[f"E{current_row}"].value = member_data['name'] if is_katowice else member_data['team_name']                       # Col E: Processor Name
                    ws[f"F{current_row}"].value = case_number                                                                            # Col F: Reference Number
                    ws[f"G{current_row}"].value = case[service_col]                                                                      # Col G: HR Service
                    # ── To add a column: ws[f"X{current_row}"].value = case[new_col]
                    #    (replace X with the QCL column letter, e.g. "H" for column H)
                    # ─────────────────────────────────────────────────────────────────────────

                    # Log statistics to terminal, unprocessed notices to GUI console
                    if recognized is True:
                        print(f"QA Checks {control_no}: Case {case_number} from {location} was taken from row {raw_row_num} in {raw_file_name}")
                    elif recognized is False:
                        log_to_console(f"QA Checks {control_no}: Case {case_number} has an unrecognized location code '{raw_location}'. Taken from row {raw_row_num} in {raw_file_name}.")
                    else:  # recognized is None (blank)
                        log_to_console(f"QA Checks {control_no}: Case {case_number} has a blank location. Taken from row {raw_row_num} in {raw_file_name}.")

                    control_no += 1
                    current_row += 1
                    total_cases_written += 1

            log_to_console(f"QA Checks completed: {total_cases_written} cases written.\n")

            qa_checks_stats = {
                'cases': total_cases_written
            }

        # ========== PROCESS QA REVIEWS - REOPENED CASES ==========
        if qa_review_file:
            try:
                control_no_start = 1  # Start from 1 for QA Reviews
                _, cases_written = process_qa_reviews_reopened_cases(
                    qa_review_file,
                    member_file,
                    wb,
                    control_no_start,
                    selected_circle,
                    log_to_console,
                    is_katowice
                )

                qa_reviews_reopened_stats = {
                    'cases': cases_written
                }
            except Exception as e:
                log_to_console(f"Warning: Could not process QA Reviews - Reopened Cases - {str(e)}")
                qa_reviews_reopened_stats = None

        # ========== PROCESS QA REVIEWS - BREACHED CASES ==========
        if qa_review_file:
            try:
                control_no_start = 1  # Start from 1 for Breached Cases
                _, cases_written = process_qa_reviews_breached_cases(
                    qa_review_file,
                    member_file,
                    wb,
                    control_no_start,
                    selected_circle,
                    log_to_console,
                    is_katowice
                )

                qa_reviews_breached_stats = {
                    'cases': cases_written
                }
            except Exception as e:
                log_to_console(f"Warning: Could not process QA Reviews - Breached Cases - {str(e)}")
                qa_reviews_breached_stats = None

        # ========== PROCESS QA REVIEWS - CSATS ==========
        if qa_review_file:
            try:
                control_no_start = 1  # Start from 1 for CSATs
                _, cases_written = process_qa_reviews_csats(
                    qa_review_file,
                    member_file,
                    wb,
                    control_no_start,
                    selected_circle,
                    log_to_console,
                    is_katowice
                )

                qa_reviews_csats_stats = {
                    'cases': cases_written
                }
            except Exception as e:
                log_to_console(f"Warning: Could not process QA Reviews - CSATs - {str(e)}")
                qa_reviews_csats_stats = None

        # ========== PROCESS QA REVIEWS - CSAT CHAT ==========
        if qa_review_file:
            try:
                control_no_start = 1  # Start from 1 for CSAT Chat
                _, cases_written = process_qa_reviews_csat_chat(
                    qa_review_file,
                    member_file,
                    wb,
                    control_no_start,
                    selected_circle_text,
                    log_to_console,
                    is_katowice
                )

                qa_reviews_csat_chat_stats = {
                    'cases': cases_written
                }
            except Exception as e:
                log_to_console(f"Warning: Could not process QA Reviews - CSAT Chat - {str(e)}")
                qa_reviews_csat_chat_stats = None

        # ========== PROCESS QA REVIEWS - LIVE CHAT FCR ==========
        if qa_review_file:
            try:
                control_no_start = 1
                _, cases_written = process_qa_reviews_live_chat_fcr(
                    qa_review_file,
                    member_file,
                    wb,
                    control_no_start,
                    selected_circle_text,
                    log_to_console,
                    is_katowice
                )

                qa_reviews_live_chat_fcr_stats = {
                    'cases': cases_written
                }
            except Exception as e:
                log_to_console(f"Warning: Could not process QA Reviews - Live Chat FCR - {str(e)}")
                qa_reviews_live_chat_fcr_stats = None

        # ========== PROCESS QA REVIEWS - WD BREACH ==========
        if qa_review_file:
            try:
                control_no_start = 1
                _, cases_written = process_qa_reviews_wd_breach(
                    qa_review_file,
                    wb,
                    control_no_start,
                    selected_circle_text,
                    log_to_console
                )

                qa_reviews_wd_breach_stats = {
                    'cases': cases_written
                }
            except Exception as e:
                log_to_console(f"Warning: Could not process QA Reviews - WD Breach - {str(e)}")
                qa_reviews_wd_breach_stats = None

        # Save workbook
        wb.save(output_path)
        wb.close()

        # Build success message
        print("=== PROCESSING COMPLETE ===")

        success_msg = "QCL generated successfully!\n\n"
        success_msg += f"Output: {output_path}\n\n"

        if qa_checks_stats:
            success_msg += f"QA Checks: {qa_checks_stats['cases']} cases written\n"

        if qa_reviews_reopened_stats:
            success_msg += f"Reopened Cases: {qa_reviews_reopened_stats['cases']} cases written\n"

        if qa_reviews_breached_stats:
            success_msg += f"Breached Cases: {qa_reviews_breached_stats['cases']} cases written\n"

        if qa_reviews_csats_stats:
            success_msg += f"CSAT: {qa_reviews_csats_stats['cases']} cases written\n"

        if qa_reviews_csat_chat_stats:
            success_msg += f"CSAT Chat: {qa_reviews_csat_chat_stats['cases']} cases written\n"

        if qa_reviews_live_chat_fcr_stats:
            success_msg += f"Live Chat FCR: {qa_reviews_live_chat_fcr_stats['cases']} cases written\n"

        if qa_reviews_wd_breach_stats:
            success_msg += f"WD Breach: {qa_reviews_wd_breach_stats['cases']} cases written"

        messagebox.showinfo("Success", success_msg)

    except Exception as e:
        messagebox.showerror("Error", str(e))
        import traceback
        traceback.print_exc()

# ---------------- GUI SETUP ---------------- #

root = tk.Tk()
root.title("AutoMaster")
root.geometry("600x700")
root.minsize(500, 600)  # Set minimum size to prevent elements from overlapping
root.configure(bg="#BFBFBF")

# Set window icon
try:
    icon_image = Image.open(get_resource_path("icon.png"))
    icon_photo = ImageTk.PhotoImage(icon_image)
    root.iconphoto(True, icon_photo)
except Exception as e:
    print(f"Could not load icon: {e}")

# Create main container with scrollbar for small screens
main_canvas = tk.Canvas(root, bg="#BFBFBF", highlightthickness=0)
scrollbar = tk.Scrollbar(root, orient="vertical", command=main_canvas.yview)
scrollable_frame = tk.Frame(main_canvas, bg="#BFBFBF")

scrollable_frame.bind(
    "<Configure>",
    lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
)

main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
main_canvas.configure(yscrollcommand=scrollbar.set)

# Pack canvas and scrollbar
main_canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

# Enable mouse wheel scrolling
def on_mousewheel(event):
    main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

main_canvas.bind_all("<MouseWheel>", on_mousewheel)

# Console frame
console_label = tk.Label(scrollable_frame, text="Console:", font=("Arial", 12, "bold"),
                         bg="#BFBFBF", fg="black")
console_label.pack(anchor=tk.W, padx=20, pady=(10, 5))

console_frame = tk.Frame(scrollable_frame, bg="white", bd=2, relief=tk.SOLID)
console_frame.pack(fill=tk.X, padx=20, pady=(0, 15))

console_text = scrolledtext.ScrolledText(console_frame, height=10, font=("Consolas", 9),
                                          bg="#1e1e1e", fg="#d4d4d4", wrap=tk.WORD,
                                          state=tk.DISABLED)
console_text.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

def log_to_console(message):
    """Log a message to the console text area"""
    console_text.config(state=tk.NORMAL)
    console_text.insert(tk.END, message + "\n")
    console_text.see(tk.END)  # Auto-scroll to bottom
    console_text.config(state=tk.DISABLED)
    console_text.update()  # Force update to show message immediately

def clear_console():
    """Clear the console text area"""
    console_text.config(state=tk.NORMAL)
    console_text.delete(1.0, tk.END)
    console_text.config(state=tk.DISABLED)

# Circle Selection
circle_label = tk.Label(scrollable_frame, text="1.", font=("Arial", 14, "bold"), 
                        bg="#BFBFBF", fg="black")
circle_label.pack(anchor=tk.W, padx=30, pady=(10, 5))

circle_frame = tk.Frame(scrollable_frame, bg="white", bd=2, relief=tk.SOLID, highlightbackground="#FF6B35", 
                        highlightthickness=2)
circle_frame.pack(fill=tk.X, padx=30, pady=5)

circle_var = tk.StringVar(value="All Circles")
circle_dropdown = tk.OptionMenu(
    circle_frame,
    circle_var,
    "All Circles",
    "Circle 1 (HCM, OM)",
    "Circle 2 (Expense, Reporting)",
    "Circle 3 (Learning)",
    "Circle 4 (International Mobility)",
    "Circle 5 (Performance & Rewards, Travel)",
    "Circle 6 (Contact Center)",
    "Circle 6 (Recruitment Admin)",
    "All Katowice",
    "HCM - BE",
    "HCM - NL",
    "People Services - DE",
    "People Services - NL",
    "People Services - PL"
)
circle_dropdown.config(font=("Arial", 11), bg="white", fg="gray", 
                       relief=tk.FLAT, highlightthickness=0, width=35)
circle_dropdown.pack(fill=tk.X, padx=5, pady=5)

# File upload section
file_section_label = tk.Label(scrollable_frame, text="2.", font=("Arial", 14, "bold"), 
                              bg="#BFBFBF", fg="black")
file_section_label.pack(anchor=tk.W, padx=30, pady=(15, 5))

file_container = tk.Frame(scrollable_frame, bg="white", bd=2, relief=tk.SOLID)
file_container.pack(fill=tk.X, padx=30, pady=5)

def create_file_row(parent, icon_text, button_text, status_text):
    row_frame = tk.Frame(parent, bg="white")
    row_frame.pack(fill=tk.X, padx=10, pady=8)
    
    icon_frame = tk.Frame(row_frame, bg="#00C853", width=40, height=40)
    icon_frame.pack(side=tk.LEFT, padx=(0, 10))
    icon_frame.pack_propagate(False)
    icon_label = tk.Label(icon_frame, text=icon_text, font=("Arial", 16, "bold"), 
                         bg="#00C853", fg="white")
    icon_label.pack(expand=True)
    
    button = tk.Button(row_frame, text=button_text, font=("Arial", 10, "bold"),
                      bg="#FF6B35", fg="white", width=18, relief=tk.FLAT, cursor="hand2")
    button.pack(side=tk.LEFT)
    
    status = tk.Label(row_frame, text=status_text, font=("Arial", 9),
                     bg="gray", fg="white", anchor=tk.W, padx=10)
    status.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))
    
    entry = tk.Entry(row_frame)
    
    def browse_file():
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm")])
        if path:
            entry.delete(0, tk.END)
            entry.insert(0, path)
            filename = path.split("/")[-1]
            status.config(text=filename, bg="green")
    
    button.config(command=browse_file)
    return entry

qa_checks_entry = create_file_row(file_container, "📊", "QA CHECKS RAW", "No raw data selected")
qa_review_entry = create_file_row(file_container, "📊", "QA REVIEW RAW", "No raw data selected")

# QCL Template
qcl_label = tk.Label(scrollable_frame, text="3.", font=("Arial", 14, "bold"), 
                     bg="#BFBFBF", fg="black")
qcl_label.pack(anchor=tk.W, padx=30, pady=(15, 5))

qcl_frame = tk.Frame(scrollable_frame, bg="white", bd=2, relief=tk.SOLID)
qcl_frame.pack(fill=tk.X, padx=30, pady=5)
qcl_entry = create_file_row(qcl_frame, "📄", "QCL TEMPLATE", "No template selected")

# Member List
member_label = tk.Label(scrollable_frame, text="4.", font=("Arial", 14, "bold"), 
                        bg="#BFBFBF", fg="black")
member_label.pack(anchor=tk.W, padx=30, pady=(15, 5))

member_frame = tk.Frame(scrollable_frame, bg="white", bd=2, relief=tk.SOLID)
member_frame.pack(fill=tk.X, padx=30, pady=5)
member_entry = create_file_row(member_frame, "👤", "MEMBER LIST", "No member list selected")

# AutoMate Button
automate_btn = tk.Button(scrollable_frame, text="AutoMate", font=("Arial", 14, "bold"),
                        bg="#FF6B35", fg="white", width=20, height=2,
                        relief=tk.FLAT, cursor="hand2", command=process_files)
automate_btn.pack(pady=30)
root.mainloop()
